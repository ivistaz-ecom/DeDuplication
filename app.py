from flask import Flask, request, render_template, jsonify, send_file, redirect, url_for, flash, send_from_directory, abort, Response, session
import os
import pickle
from pymongo import MongoClient, ASCENDING, DESCENDING
import certifi
import shutil
from datetime import datetime
import pandas as pd
import chardet
import encodings
from io import BytesIO
import numpy as np
import re
from threading import Thread
import Levenshtein
import ast
import time
import demoji
import unicodedata
import openai
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import io
import warnings
from functools import wraps
# import csv
# from langdetect.lang_detect_exception import LangDetectException
# import fasttext
# from langdetect import detect, LangDetectException
#from googletrans import Translator, LANGUAGES
#start openai api lib
#end openai api lib
warnings.simplefilter(action='ignore', category=FutureWarning)

app = Flask(__name__)
app.config['FILE_STORE'] = './file_store' #used to store files permnantely....

app.secret_key = 'SeoDashboard' #secret key for session creation
UPLOAD_FOLDER = "./uploads"  # temporary folder


uploaded_data = pd.DataFrame() #holds the data uploaded at fileupload 
global_uploaded_data = pd.DataFrame() # used to store the deep copy of uploaded data for only .csv files
state_and_city_dictionary = {} #holds the state:[city] dictionary for each state
duplicate_data = pd.DataFrame() #holds the column names of uploaded data as dataframe
column_headings = []  # holds list of column names
city_list = []   # holds the city list in state_and_city.csv
email_domain_combined_data = {} # stores correct domain keys and incorrect domain values as list Eg: 'gmail.com':['gmil.com','gmail.co']
translation_count=0 # counting translation  one language to another
tasks = {} #handling tasks  
global_duplicated_data_all=pd.DataFrame() #used to store duplicated data into dataframe
ExportDataEditOkOriginal = pd.DataFrame() #used to storing data at selected export file for editing and also stores data after filtering
ExportDataEditOkNew = pd.DataFrame()  #used to holds the coulmns of selected file as dataframe and also filtered data(invalid data)
ExportDataEditUploadFileData = pd.DataFrame()#used to store the file uploaded data at Export data
ReportDataGlobal = pd.DataFrame() #used to store the report data as dataframe


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        print("Checking authentication")
        print(session.get('is_authenticated'))
        if not session.get("is_authenticated"):
            flash("You must log in to access this page.")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated_function


# index page...
@app.route('/')
def index():
    return render_template("pages/index_first_page.html")

# making database connection then checking credentials and making session for user
@app.route('/dashboard', methods=['POST', 'GET'])
def dashboard():    
    email = request.form.get('email')
    password = request.form.get('password')
    if 'user' in session:
        return redirect(url_for('showDashboard'))  # sesssion handled
    else:
        try:

            if os.path.exists("./database_connection/database_connection.pkl"):
                directory = './database_connection'
                file_path = os.path.join(directory, 'database_connection.pkl')
                if os.path.exists(file_path):
                    with open(file_path, "rb") as f:
                        conn_data = pickle.load(f)
                else:
                    conn_data = {}
            client = MongoClient(
                conn_data['database_connection'],
                tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
            collection = client['userDetail']
            """
            client = MongoClient(
                "mongodb+srv://anithadevi:AnithaDevi02011998@cluster0.uyhgg.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0",
                tlsCAFile=certifi.where()).get_database('SeoDataDatabase')
            collection = client['userDetail']
            """
            query = {
                '$and': [
                    {"email": email},
                    {"password": password}
                ]
            }
            document = collection.find_one(query)
            # print("document:",document)   
            if document is not None:
                # making session
                session["is_authenticated"] = True
                session['user'] = email
                session['name'] = document['name']
                session['usertype'] = document['usertype']
                return redirect(url_for('showDashboard'))
            return redirect(url_for('index'))
        except Exception as e:
            return jsonify({"status": "error", "message": f"{e}"}), 500

"""
                                                    DashBoard Starts

"""
# showdashboard Page

@app.route('/showDashboard')
@login_required
def showDashboard():
    # return render_template("pages/index.html")
    return render_template("components/dashboard_first_page.html")


#Sending the name you see in after Clicking PrepareData Button
@app.route('/getDatabaseCollection', methods=['POST'])
@login_required
def getDatabaseCollection():
    conn_data = {}
    if os.path.exists("./database_connection/database_connection.pkl"):
        directory = './database_connection'
        file_path = os.path.join(directory, 'database_connection.pkl')
        if os.path.exists(file_path):
            with open(file_path, "rb") as f:
                conn_data = pickle.load(f)
        else:
            conn_data = {}
    if len(conn_data) > 0:
        result = conn_data['collection_name']
        result = " ".join(result.split("_"))
        result = conn_data['source_type'] + " " + result
        result = result.lstrip(" ")
        result = result.capitalize()
        return jsonify(result=result)
    else:
        result = "None"
        return jsonify(result=result)
    # return Response(status=204)

#prepare the  Data based on provided ids fetching data from Mongodb and Storing that file in download_database folder

@app.route('/prepareDatabaseData', methods=['POST'])
@login_required
def prepareDatabaseData():
    try:
        # new added
        data = request.json
        column_name = data.get('columnName')
        field1_value = int(data.get('field1'))
        field2_value = int(data.get('field2'))
        # new added end

        conn_data = {}
        # print("database prepare testing:")
        """
        p=True
        if p:
            raise FileNotFoundError("database_connection.pkl file not found.")
        """

        if os.path.exists("./download_database"):
            shutil.rmtree("./download_database")

        if os.path.exists("./database_connection/database_connection.pkl"):
            directory = './database_connection'
            file_path = os.path.join(directory, 'database_connection.pkl')
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    conn_data = pickle.load(f)
        if len(conn_data) > 0:
            client = MongoClient(
                conn_data['database_connection'],
                tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
            # collection = client[conn_data['collection_name']]
            coll_name_temp = conn_data['source_type'] + "_" + conn_data['collection_name']
            coll_name_temp = coll_name_temp.lstrip("_")
            collection = client[coll_name_temp]
            if (field1_value == 0 or field2_value == 0 or field1_value > field2_value):
                cursor = collection.find({}, {"_id": 0})
                count_documents = collection.count_documents({})
                if count_documents > 0:
                    data = list(cursor)
                    # Create an in-memory output buffer
                    dataframe = pd.DataFrame(data)
                    try:
                        if not os.path.exists("./download_database"):
                            os.makedirs("./download_database")
                        # dataframe.to_csv(os.path.join("./download_database","walmart_vriddhi_database.csv"),index=False)
                        dataframe.to_csv(os.path.join("./download_database", f"{coll_name_temp}.csv"), index=False)
                        return jsonify({"status": "success"})
                    except Exception as e:
                        return jsonify({"status": "error", "message": f"{e}"}), 500
                else:
                    return jsonify({"status": "error", "message": "Database is empty!"}), 500
                # print("prepare data completed")
                # return jsonify({"status": "success"})
            else:
                try:
                    query = {
                        column_name: {
                            '$gte': field1_value,  # Greater than or equal to id_min
                            '$lte': field2_value  # Less than or equal to id_max
                        }
                    }
                    results = collection.find(query, {'_id': 0}) # _id:0 ignore the _id in mongodb
                    t_data = list(results)
                    if len(t_data) > 0:
                        dataframe = pd.DataFrame(t_data)
                        try:
                            if not os.path.exists("./download_database"):
                                os.makedirs("./download_database")
                            dataframe.to_csv(os.path.join("./download_database", f"{coll_name_temp}.csv"), index=False)
                            return jsonify({"status": "success"})
                        except Exception as e:
                            return jsonify({"status": "error", "message": f"{e}"}), 500
                    else:
                        return jsonify({"status": "error", "message": "Database is empty!"}), 500
                except Exception as e:
                    return jsonify({"status": "error", "message": f"{e}"}), 500
        else:
            return jsonify({"status": "error", "message": "There is no database connection setting!"}), 500
            # return Response(status=204)
    except Exception as e:
        # print("exception occured as ")
        return jsonify({"status": "error", "message": str(e)}), 500

#Showing database Data
@app.route('/showDatabase', methods=['POST'])
@login_required
def showDatabase():
    conn_data = {}
    client = None
    try:

        button_id = request.args.get('id')
        # data_table=pd.DataFrame()
        if os.path.exists("./database_connection/database_connection.pkl"):
            directory = './database_connection'
            file_path = os.path.join(directory, 'database_connection.pkl')
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    conn_data = pickle.load(f)

                # database connection
            try:
                client = MongoClient(
                    conn_data['database_connection'],
                    tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
            except Exception as e:
                return jsonify({"status": "error", "message": "in show database section" + str(e)}), 500
        #ID 1 belongs to Business catalyst
        if button_id == '1':
            if client != None:
                collection = client["business_catalyst"]
                documents = list(collection.find({}, {"_id": 0}).limit(1000000))
                if len(documents) > 0:
                    df2 = pd.DataFrame(documents)
                    df2['id'] = df2['id'].astype(int)
                    data_table = df2.to_html(classes='dataframe', index=False, escape=False)
                    return jsonify({"status": "success", "data_table": data_table}), 200
                else:
                    return jsonify({"status": "error", "message": "There is not data in database"}), 500
        # Both 2 & 3 Ids belongs to Walmart Vriddhi(Website & Paid)
        if button_id == '2':

            try:
                collection = client["Website_walmart_vriddhi"]
                cursor = collection.find({}, {"_id": 0}).sort('id', DESCENDING).limit(10000)
                try:
                    documents = list(cursor)
                    if len(documents) > 0:
                        df2 = pd.DataFrame(documents) #converting data from database into DataFrame
                        df2['id'] = df2['id'].astype(int) #converting data type of id to {int}
                        data_table = df2.to_html(classes='dataframe', index=False, escape=False)
                        return jsonify({"status": "success", "data_table": data_table}), 200
                    else:
                        return jsonify({"status": "error", "message": "There is not data in database"}), 500
                except Exception as e:
                    return jsonify({"status": "error", "message": str(e)}), 500
            except Exception as e:
                return jsonify({"status": "error", "message": "in database connection" + str(e)}), 500
        if button_id == '3':
            try:
                collection = client["Paid_walmart_vriddhi"]
                documents = list(collection.find({}, {"_id": 0}).limit(100000))
                if len(documents) > 0:
                    df2 = pd.DataFrame(documents)
                    df2['id'] = df2['id'].astype(int)
                    data_table = df2.to_html(classes='dataframe', index=False, escape=False)
                    return jsonify({"status": "success", "data_table": data_table}), 200
                else:
                    return jsonify({"status": "error", "message": "There is not data in database"}), 500
            except Exception as e:
                return jsonify({"status": "error", "message": "Error in Walmart Market place"}), 500
        #Both 4 & 5 Ids belongs to walmart market Place (Website & Paid)
        if button_id == '4':

            try:
                collection = client["Website_walmart_market_place"]
                cursor = collection.find({}, {"_id": 0}).sort('id', DESCENDING).limit(10000)
                try:
                    documents = list(cursor)
                    if len(documents) > 0:
                        df2 = pd.DataFrame(documents)
                        df2['id'] = df2['id'].astype(int)
                        data_table = df2.to_html(classes='dataframe', index=False, escape=False)
                        return jsonify({"status": "success", "data_table": data_table}), 200
                    else:
                        return jsonify({"status": "error", "message": "There is not data in database"}), 500
                except Exception as e:
                    return jsonify({"status": "error", "message": str(e)}), 500
            except Exception as e:
                return jsonify({"status": "error", "message": "in database connection" + str(e)}), 500
        if button_id == '5':
            try:
                collection = client["Paid_walmart_market_place"]
                documents = list(collection.find({}, {"_id": 0}).limit(10000))
                if len(documents) > 0:
                    df2 = pd.DataFrame(documents)
                    df2['id'] = df2['id'].astype(int)
                    data_table = df2.to_html(classes='dataframe', index=False, escape=False)
                    return jsonify({"status": "success", "data_table": data_table}), 200
                else:
                    return jsonify({"status": "error", "message": "There is not data in database"}), 500
            except Exception as e:
                return jsonify({"status": "error", "message": "Error in Walmart Market place"}), 500

    except Exception as e:
        # Handle the error and return a JSON response with an error status
        # print(f"exception: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

#Download the database
@app.route('/downloadDatabaseData', methods=['POST'])
@login_required
def downloadDatabaseData():
    # print("database download testing:")
    if os.path.exists('./download_database'):
        list_of_files = list(os.listdir('./download_database'))
        if len(list_of_files) > 0:
            file_path = os.path.join('./download_database', list_of_files[0])
            try:
                return send_file(file_path, as_attachment=True)
            except Exception as e:
                return jsonify({"status": "error", "message": str(e)}), 500
                # abort(404)  # Return 404 if file not found
        else:
            return Response(status=204)
    else:
        return Response(status=204)

"""
                                                    Dashboard Ends
                                                    Deduplication Starts
"""
# sending the database info if connection established correctly...
#start Deduplication
@app.route('/databaseConnection')
@login_required
def databaseConnection():

    directory = './database_connection'
    file_path = os.path.join(directory, 'database_connection.pkl')

    try:
        if os.path.exists(file_path):
            with open(file_path, "rb") as f:
                data = pickle.load(f)

        else:
            data = {}
        # print("complete database connection data: ",data)
        # remove already existing folder/file current_data folder used in Datapreprocessing Page
        if os.path.exists("./current_data"):
            shutil.rmtree("./current_data")
        # removing already existing folder ExportDataEditFile used in Export Page
        folder_path = "./ExportDataEditFile"
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)
        # removing already existing folder ExportDataEditUploadFile used in Export Page
        folder_path = "./ExportDataEditUploadFile"
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)

    except Exception as e:
        app.logger.error(f"Error in /databaseConnection: {e}")
        return "Internal Server Error", 500

    return render_template("components/database_connection.html", data=data, usertype=session['usertype'])


# Editing the database collections and source_type

@app.route('/database_edit', methods=["POST"])
@login_required
def database_edit():
    # Get the data from the form
    database_connection = request.form.get('connection_string')
    database_name = request.form.get('database_name')
    collection_name = request.form.get('collection_name')
    source_type = request.form.get('source_type_name')
    # if collection_name=="business_catalyst":
    source_type = "" if collection_name == "business_catalyst" else source_type
    #print("source_type: ", source_type, collection_name)
    # print("source_type: ",source_type)
    list_of_collections = ['walmart_vriddhi', 'business_catalyst', 'walmart_market_place']
    list_source_type = ['Paid', 'Website']
    data = {"database_connection": database_connection, "database_name": database_name,
            "collection_name": collection_name, "source_type": source_type}
    directory = './database_connection'
    if not os.path.exists(directory):
        os.makedirs(directory)
    file_path = os.path.join(directory, 'database_connection.pkl')
    try:
        """
        if not os.path.exists(directory):
            os.makedirs(directory)
        """
        if collection_name == "business_catalyst":
            with open(file_path, "wb") as f:
                pickle.dump(data, f)
            flash('Edit completed successfully!', 'success')
        elif data["collection_name"] not in list_of_collections or data["source_type"] not in list_source_type:
            flash('Please add options based on available databases!', 'error')
        else:
            with open(file_path, "wb") as f:
                pickle.dump(data, f)
            flash('Edit completed successfully!', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'error')
    return redirect(url_for("databaseConnection"))

# character encoding for csv files.... chardet returns a dictionary = {"encoding":"encoding_type-utf,ascii...","confidence":0.99}
def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']

"""
                                                Start of File Uploading
"""
@app.route('/fileupload', methods=["POST", "GET"])
@login_required
def fileupload():
    status = True
    uploaded_file = request.files.getlist("file_input")
    data_separator_symbol = request.form.get('text_input')

    for file in uploaded_file:
        if file.filename == "":
            flash("Not valid file, will not be uploaded!", "error")
            return redirect(url_for("fileupload"))

        #used to store uploaded files temporary in "uploads folder"
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)

        # By using data and time changing the uploaded file name Eg: base_20241217_112233.csv
        now = datetime.datetime.now()
        date_str = now.strftime("%Y%m%d")  # Format as YYYYMMDD
        time_str = now.strftime("%H%M%S")  # Format as HHMMSS
        original_filename = file.filename
        base, extension = os.path.splitext(original_filename)
        new_filename = f"{base}_{date_str}_{time_str}{extension}"  # YYYYMMDD-HHMMSS

        data_separator_symbol = None if (
            data_separator_symbol == None or data_separator_symbol == '') and extension == ".csv" else data_separator_symbol

        global uploaded_data
        if extension == ".xlsx":
            file.save(f"{UPLOAD_FOLDER}/" + file.filename) #saving file on ./uploads folder
            file_path = f"./uploads/{file.filename}"
            new_file_path = f"./uploads/{new_filename}"

            # rename file with datetime
            os.rename(file_path, new_file_path)

            #uploading data into dataframe 
            uploaded_data = pd.read_excel(new_file_path)
            flash('File uploaded successfully!', 'success')
            return redirect(url_for("fileupload"))

        elif extension == ".csv":
            file.save(f"{UPLOAD_FOLDER}/" + file.filename)
            file_path = f"./uploads/{file.filename}"
            new_file_path = f"./uploads/{new_filename}"

            # rename file with datetime
            os.rename(file_path, new_file_path)
            # sending the file path to encoding function it results what type of enconding used in this
            file_decode = detect_encoding(new_file_path)
            try:
                uploaded_data = pd.read_csv(new_file_path, encoding=file_decode, sep=data_separator_symbol, engine="python")
            except Exception as e:
                try:
                    uploaded_data = pd.read_csv(new_file_path, sep=data_separator_symbol, engine="python")
                except Exception as e1:
                    status = (not status)
                    if file_decode not in list(set(encodings.aliases.aliases.values())):
                        flash(f'{file_decode} encoding format of file not existing in data encoding list!', 'error')
                    else:
                        flash(f'{e1}', 'error')

            global global_uploaded_data
            global_uploaded_data = uploaded_data.copy(deep=True)
            if status:
                flash('File uploaded successfully!', 'success')
            return redirect(url_for("fileupload"))
        else:
            flash("Wrong file format to upload, File will not be uploaded!", "error")
    return render_template("components/file_upload.html")

#showing uploaded data as table in the same page
@app.route('/showdata')
@login_required
def show_data():
    global uploaded_data
    # Check if uploaded_data is empty
    if len(uploaded_data) == 0:
        # If current_data.csv exists, load it into uploaded_data (The data changed at data Preprocessing stored in current_data.csv file)
        if os.path.exists("./current_data/current_data.csv"):
            #placing content in current_data.csv file to uploaded_data dataframe
            uploaded_data = pd.read_csv("./current_data/current_data.csv")

    # Check again if uploaded_data is empty after attempting to load data
    if len(uploaded_data) == 0:
        return jsonify({'data_table': '', 'len_is_zero': True})  # Return empty table and flag

    # Convert DataFrame to HTML table
    data_table = uploaded_data.to_html(classes='dataframe', index=False, escape=False)
    return jsonify({'data_table': data_table, 'len_is_zero': False})  # Include flag for len_is_zero

# downlaod the csv file converting df 
@app.route('/download/csv', methods=['GET'])
@login_required
def download_csv():
    buffer = BytesIO()
    # Write DataFrame to CSV
    uploaded_data.to_csv(buffer, index=False)
    # Making File Pointer at the Begining of the file
    buffer.seek(0)
    # Return the CSV file
    return send_file(buffer, as_attachment=True, download_name='uploaded_data.csv', mimetype='text/csv')

@app.route('/download/xlsx', methods=['GET'])
@login_required
def download_xlsx():
    buffer = BytesIO()
    # Write DataFrame to Excel
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        uploaded_data.to_excel(writer, index=False, sheet_name='Sheet1')
    buffer.seek(0)
    # Return the Excel file
    return send_file(buffer, as_attachment=True, download_name='uploaded_data.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

"""                                             
                                               End Of FileUpload
                                               Data Preprocessing Started
"""

@app.route('/dataPreprocessing', methods=['POST', 'GET'])
@login_required
def dataPreprocessing():
    try:
        # removing all files from upload folder to file store folder
        src_folder = "./uploads"
        dst_folder = "./file_store"
        if not os.path.exists(dst_folder):
            os.makedirs(dst_folder)  # Create the destination folder if it does not exist

        # create folder to save the current data
        if not os.path.exists("./current_data"):
            os.makedirs("./current_data")  # Create the destination folder if it does not exist

            # Copy all files and folders from src_folder to dst_folder
        if os.path.exists(src_folder):
            for item in os.listdir(src_folder):
                #combine items of source folder to paths
                src_path = os.path.join(src_folder, item)
                dst_path = os.path.join(dst_folder, item)
                if os.path.isdir(src_path):
                    # Recursively copy subdirectories if destination folder exists overwrites it
                    shutil.copytree(src_path, dst_path, dirs_exist_ok=True)
                else:
                    # Copy files including metadata
                    shutil.copy2(src_path, dst_path)

            # Remove the source folder(Removing uploads folder)
            shutil.rmtree(src_folder)
        
        # cities list handling
        state_city_data = pd.read_csv("./state_and_city_list/wvcities.csv") # state& city dataframe
        grouped = state_city_data.groupby('State')['City'].apply(list) # grouping = andhra [krishna,bapatla...]
        state_city_dict = grouped.to_dict() #converting dataframe into dictionary

        # handling state and city globally
        global state_and_city_dictionary
        state_and_city_dictionary = state_city_dict #storing same dict to global

        # handling duplicate data
        global duplicate_data
        duplicate_data = pd.DataFrame(columns=uploaded_data.columns)
        # handling columns of dataframe
        column_heading = []
        if not uploaded_data.empty:
            column_heading = uploaded_data.columns.tolist() #converting dataframe columns into a list

        global column_headings
        column_headings = column_heading

        global city_list
        city_list = list(state_city_data['City'])
        column_headings.append("None")
        return render_template("components/dataPreprocessing.html", column_headings=column_headings,
                               state_city_dict=state_and_city_dictionary, city_list=city_list)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error: {e}"}), 500


# handle text replacement
@app.route('/text_replacement', methods=["POST"])
@login_required
def text_replacement():
    """
    selected_column= request.form.get('dropdown1')
    old_text= request.form.get('replacement_input1')
    new_text= request.form.get('replacement_input2')
    """
    # print("text_replacement called!")
    data = request.get_json()
    old_text = data.get('replacement_input1')
    new_text = data.get('replacement_input2')
    selected_column = data.get('dropdown1')
    # print(selected_column,old_text,new_text)
    # time.sleep(5)
    try:
        if selected_column != "Select option" and selected_column != None and old_text != "" and new_text != "":
            selected_column=selected_column.strip() #handle extra spaces
            uploaded_data[selected_column] = uploaded_data[selected_column].str.replace(old_text.strip(), new_text.strip())
        else:
            return jsonify({"status": "error", "message": "Wrong input data!"})
        # save the current data:
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}!"})
    # return Response(status=204)
    # save the current data:
    # uploaded_data.to_csv("./current_data/current_data.csv",index=False)
    # return jsonify({"status": "success"})


# handle Handle Null Value
@app.route('/handle_null_value', methods=["POST"])
@login_required
def handle_null_value():
    
    data = request.get_json()
    selected_column = data.get('dropdown2')
    selected_method = data.get('dropdown_null_handle')
    custom_na_value = data.get('handle_null_input1')
    custom_fill_value = data.get('handle_null_input2')

    global uploaded_data
    try:
        if selected_column != "Select option" and selected_column != "None":
            selected_column=selected_column.strip()
            #used for misssing values
            if selected_method == "Remove":
                #used to remove empty rows
                if custom_na_value == "":
                    uploaded_data = uploaded_data[uploaded_data[selected_column].notna()]
                # used to remove specified data like NA,N/A,missnng like this....
                else:
                    uploaded_data = uploaded_data[~uploaded_data[selected_column].isin([x.strip() for x in custom_na_value.split(",")])]
            #used for filling the missing values
            elif selected_method == "Fill":
                # if column has specified data like NA,N/A,.... firstly it replaced with Not A Number(nan) then fill with provided value
                if custom_na_value != "":
                    if custom_fill_value != "":
                        uploaded_data[selected_column] = uploaded_data[selected_column].replace(
                            [x.strip() for x in custom_na_value.split(",")], np.nan)
                        uploaded_data[selected_column] = uploaded_data[selected_column].fillna(custom_fill_value.strip())
                    else:
                        # print("enter valid input")
                        pass
                # if column has empty cell then fill it with provided  value....
                else:
                    if custom_fill_value != "":
                        uploaded_data[selected_column] = uploaded_data[selected_column].fillna(custom_fill_value.strip())
                    else:
                        # print("enter valid input")
                        pass
        else:
            return jsonify({"status": "error", "message": "Please select correct column!"})
        # save the current data:
        uploaded_data.to_csv("./current_data/current_data.csv",index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}!"})

    # return Response(status=204)
    # save the current data:
    # uploaded_data.to_csv("./current_data/current_data.csv",index=False)
    # return jsonify({"status": "success"})


# start handle alter column position and import current date
@app.route('/alterColumnPosImportDate', methods=["POST"])
@login_required
def alterColumnPosImportDate():
    """
    selected_column= request.form.get('dropdown_alter_column')
    selected_method= request.form.get('dropdown_alter_column_date')
    ordered_columns= request.form.get('dropdown_alter_column1')
    date_column_name=request.form.get('dropdown_alter_column_date2')
    """
    data = request.get_json()
    selected_column = data.get('dropdown_alter_column')
    selected_method = data.get('dropdown_alter_column_date')
    ordered_columns = data.get('dropdown_alter_column1')  # comma separated column names as string Eg:name,email,phone....
    date_column_name = data.get('dropdown_alter_column_date2') #the name given to date column in uploaded data

    global uploaded_data
    try:
        if selected_method == "Alter columns pos":
            # alter the column position
            temp = [x.strip(" ") for x in ordered_columns.split(",")] # list of column names you have altered 
            if ("Select option" not in temp) and ("None" not in temp):
                uploaded_data = uploaded_data[[x.strip(" ") for x in ordered_columns.split(",")]]
            else:
                return jsonify({"status": "error", "message": f"Please select correct columns name!"})
        elif selected_method == "Import date":
            # import date column added
            if len(date_column_name) > 0 and (date_column_name not in uploaded_data.columns.tolist()):
                uploaded_data[date_column_name] = str(datetime.datetime.now().date())
            else:
                return jsonify({"status": "error", "message": f"Something wrong with name of date column!"})
        # print("selected_column:",selected_column,"selected_method:",selected_method,"ordered_columns:",ordered_columns.split(','),"date_column_name:",date_column_name)
        # return Response(status=204)

        # save the current data:
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}!"})


# end handle alter column position and import current date

# handle mobile no
@app.route('/handle_mobile_no', methods=["POST"])
@login_required
def handle_mobile_no():
    """
    selected_column= request.form.get('dropdown3')
    prefix= request.form.get('mobile_input1')
    suffix= request.form.get('mobile_input2')
    middle_value=request.form.get('mobile_input3')
    """
    data = request.get_json()
    selected_column = data.get('dropdown3')
    prefix = data.get('mobile_input1')
    suffix = data.get('mobile_input2')
    middle_value = data.get('mobile_input3')
    try:
        if len(selected_column) > 0 and selected_column != 'Select option' and selected_column != "None":
            selected_column=selected_column.strip() #handle extra space
            uploaded_data[selected_column] = uploaded_data[selected_column].astype(str)
            # Replace starting substring with empty string
            if len(prefix) > 0:
                uploaded_data[selected_column] = uploaded_data[selected_column].apply(
                    lambda x: x.replace(prefix, '', 1) if (x.startswith(prefix) and (len(x) - len(prefix)) >= 10) else x)
            # replace ending substring with empty string

            if len(suffix) > 0:
                uploaded_data[selected_column] = uploaded_data[selected_column].apply(
                    lambda x: x[:-len(suffix)] if x.endswith(suffix) else x)

            # replace middle substring with empty string
            if len(middle_value) > 0:
                uploaded_data[selected_column] = uploaded_data[selected_column].str.replace(
                    r'\b' + re.escape(middle_value) + r'\b', '', regex=True)
        else:
            return jsonify({"status": "error", "message": f"Please select correct column!"})
        # return Response(status=204)
        # save the current data:
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}!"})

# start handle email
@app.route('/handle_email', methods=['POST'])
@login_required
def handle_email():
    # Get JSON data from the request
    data = request.get_json()
    selected_column = data.get('dropdown4')
    email_domain_value = data.get('email_domain_input')
    global email_domain_combined_data # a dictionary
    try:
        if email_domain_value != None and email_domain_value != "" and selected_column != "Select option" and selected_column != "None":
            email_domain_value = [x.strip() for x in email_domain_value.split(",")]
            # adding key and value as list
            for p in email_domain_value: 
                email_domain_combined_data[p] = [] # making domain keys with empty list as value eg: 'gmail.com':[]
            #print("email_domain_combined_data:",email_domain_combined_data)
            uploaded_data[selected_column] = uploaded_data[selected_column].astype(str)
            email_uploaded_list = [] # used to store domains in uploaded_data
            for x in uploaded_data[selected_column]:
                email_uploaded_list.append(x.split("@")[1])
            for x in email_uploaded_list:
                nearest = find_nearest_string(x, email_domain_value)
                email_domain_combined_data[nearest].append(x) 
            #Removing dupliactes values if present
            for x in list(email_domain_combined_data.keys()):
                email_domain_combined_data[x] = list(set(email_domain_combined_data[x]))  # key as right domain, values as wrong domain
        else:
            # return jsonify({"status": "error", "message": "Please select valid column!"}), 500
            # return jsonify("")
            pass
        # save current uploaded_data
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        return jsonify({"status": "success", "dataset": str(email_domain_combined_data)})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    # save current uploaded_data
    # uploaded_data.to_csv("./current_data/current_data.csv", index=False)

    # return jsonify({"status":"success","dataset":str(email_domain_combined_data)})

# algorithm
def find_nearest_string(target_string, strings_list):
    nearest_string = None
    min_distance = float('inf')
    for s in strings_list:
        distance = Levenshtein.distance(target_string, s)
        if distance < min_distance:
            min_distance = distance
            nearest_string = s
    return nearest_string

@app.route('/replace_email', methods=['POST'])
@login_required
def replace_email():
    data = request.get_json()
    selected_column = data.get('dropdown')
    email_domain_value = data.get('comments')
    #removing the string qutations " "
    email_domain_value = email_domain_value.strip('"')
    try:
        try:
            email_domain_value = ast.literal_eval(email_domain_value) #used to evaluate strings that contain python literal
        except (ValueError, SyntaxError) as e:
            # print(f"Error parsing string to dictionary: {e}")
            return jsonify({"status": "error", "message": str(e)}), 500
        for key in list(email_domain_value.keys()):
            for value in email_domain_value[key]:
                uploaded_data[selected_column] = uploaded_data[selected_column].apply(
                    lambda x: re.sub(r'@' + re.escape(value) + r'\b', '@' + key, x) if isinstance(x, str) else x
                )
        
        # save current uploaded_data
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        # return Response(status=204)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/handle_emoji_translation', methods=["POST"])
@login_required
def handle_emoji_translation():
    #print("Start handling translation emoji!")
    global translation_count
    translation_count = 0
    data = request.get_json()
    selected_column = data.get('dropdown5')
    emoji_status = data.get('dropdown_emoji')
    translation_status = data.get('dropdown_translation')
    language_name = data.get('dropdown_language')

    try:
        if len(selected_column) > 0 and selected_column != "Select option" and selected_column != "None":
            task_id = f'task_{int(time.time())}'  # Unique task ID based on current timestamp
            tasks[task_id] = 'Processing'  # Set initial task status

            # Define the task in a separate thread
            def long_task():
                if emoji_status == 'Emoji':
                    remove_emoji(selected_column,task_id)
                if emoji_status == 'Invalid Chars':
                    removeInvalidChars(selected_column,task_id)
                if emoji_status == 'Emoji + Invalid Chars':
                    remove_emoji(selected_column,task_id)
                    removeInvalidChars(selected_column,task_id)
                if translation_status == "Yes" and language_name != "Select option":
                    language_convert(selected_column, language_name, task_id)
                #after completing every task it should show as Complete
                tasks[task_id] = 'Complete'  # mark as complete
                # Save the current data
                if not os.path.exists("./current_data"):
                    os.makedirs("./current_data")
                uploaded_data.to_csv("./current_data/current_data.csv", index=False)
                #print("End handling translation emoji!")

            # Start the task asynchronously
            Thread(target=long_task).start()

            # Return task_id to client to track progress
            return jsonify({"status": "success", "task_id": task_id})
        else:
            return jsonify({"status": "error", "message": "Please select correct column!"})
    except Exception as e:
        error_message = str(e)
        return jsonify({"status": "error", "message": f"Error occurred: {e}!"})


# handle emoji and translation
def remove_emoji(column_name,task_id):
    #print("start removal emoji!")
    final_list = []
    for x in uploaded_data[column_name]:
        is_nan = pd.isna(x)
        if is_nan:
            final_list.append("")
        else:
            if len(x) > 1:
                # print("current data",x)
                p_t = demoji.replace(x, "")  # remove emoji
                final_list.append(p_t)
            else:
                final_list.append(x)
                # print("p_t",p_t)
    # print("final_list: ",len(final_list))
    uploaded_data[column_name] = final_list
    #tasks[task_id] = 'Complete' #mark as complete
    #print("end removal emoji!")

def removeInvalidChars(column_name,task_id):
    #print("start removal InvalidCharacters!")
    list_of_string = []
    for x in uploaded_data[column_name]:
        str1 = ""
        list1 = ["Ll", "Lu", "Nd", "Zs"]
        for char in x:
            if unicodedata.category(char) in list1:
                str1 += char
            else:
                str1 += " "
        t_data = str1.split()
        t_data = " ".join([x for x in t_data if len(x) > 0])
        list_of_string.append(t_data)
    uploaded_data[column_name] = list_of_string
    #tasks[task_id] = 'Complete' #mark as completed
    #print("end removal InvalidCharacters!")


# translator
def language_convert(column_name, language_given, task_id):
    try:
        if os.path.exists("./database_connection/database_connection.pkl"):
            with open("./database_connection/database_connection.pkl", "rb") as f:
                conn_data = pickle.load(f)
            client = MongoClient(
                conn_data['database_connection'],
                tlsCAFile=certifi.where(), tls=True)
            db = client[conn_data['database_name']]
            collection = db['api_key']
            data = collection.find_one({"id": '1'})
            if data is not None:
                openai.api_key=data['api_key']
                
        complete_translated_list = []
        all_data = [str(x) for x in uploaded_data[column_name]]
        global translation_count
        for x in all_data:
            translation_count += 1 # increasing the translation count 
            translation = translate_with_entity_preservation(x, language_given) # control goes to the below function
            complete_translated_list.append(translation)
        uploaded_data[column_name] = complete_translated_list
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"})

#Translation using openai api
def translate_with_entity_preservation(text, target_language="English"):
    # Craft the prompt to ask for translation with specific instructions on preserving entities
    prompt = (
        f"Translate the following text to {target_language}, and only provide the translation. "
        f"Keep any names or proper nouns in their original form without any additional notes or symbols.\n\n"
        f"Text: {text}"
    )

    # Make the request to OpenAI's API using a chat-based model
    response = openai.ChatCompletion.create(
        model="gpt-4o-mini-2024-07-18",
        messages=[
            {"role": "system", "content": "You are a professional translator."},
            {"role": "user", "content": prompt}
        ]
    )

    # Extract the translated text
    translated_text = response['choices'][0]['message']['content'].strip()
    return translated_text

#sending Translation count to frontend
@app.route('/check_task_status/<task_id>', methods=["GET"])
@login_required
def check_task_status(task_id):
    # Return the current status of the task
    status = tasks.get(task_id, 'Unknown')
    global translation_count
    return jsonify({"status": status,'translated_count': translation_count})

#Translation using google library
"""
def translate_to_other(text_to_translate, target_language="English"):
    lang_dict = {"English": "en", "Hindi": "hi"}
    translator = Translator()
    try:
        translation = translator.translate(text_to_translate, dest=lang_dict[target_language])
        translated_text = translation.text
        return translated_text
    except Exception as e:
        # Handle any exceptions silently
        pass
    return text_to_translate


def language_convert(column_name, language_given, task_id):
    complete_translated_list = []
    all_data = [str(x) for x in uploaded_data[column_name]]
    global translation_count
    for x in all_data:
        #print("index: ", i, ",data to translate: ", x)
        translation_count += 1
        temp_p = translate_to_other(x, language_given)
        #print("index: ", translation_count, ",data to translate: ", x, "translated text: ",temp_p)
        if temp_p is not None:
            complete_translated_list.append(temp_p)
        else:
            complete_translated_list.append(x)

    uploaded_data[column_name] = complete_translated_list
    #tasks[task_id] = 'Complete'  # Mark the task as complete after translation
"""

#handling column rename,data type conversion & date parsing..
@app.route('/handle_rename_type_conversion_date_parsing', methods=["POST"])
@login_required
def handle_rename_type_conversion_date_parsing():
    """
    selected_column= request.form.get('dropdown6')
    method_name= request.form.get('dropdown_rename')
    datatype_name= request.form.get('dropdown_type_convert')
    rename_value=request.form.get('rename_input1')
    """
    data = request.get_json()
    selected_column = data.get('dropdown6')
    method_name = data.get('dropdown_rename')
    datatype_name = data.get('dropdown_type_convert')
    rename_value = data.get('rename_input1')

    # print("selected_column",selected_column,"method_name",method_name,"datatype_name",datatype_name,"rename_value",rename_value)
    global uploaded_data
    try:
        if len(selected_column) > 0 and selected_column != "Select option" and selected_column != "None":
            selected_column=selected_column.strip()
            if method_name == "Rename":
                if rename_value != "" and (rename_value not in uploaded_data.columns.tolist()):
                    # uploaded_data = uploaded_data.rename(columns={selected_column: rename_value})
                    uploaded_data.rename(columns={selected_column: rename_value}, inplace=True)
                    # print(uploaded_data.columns)
                else:
                    # print("something wrong!")
                    return jsonify({"status": "error", "message": f"Please enter valid column name!"})
            elif method_name == "Type convert":
                if len(datatype_name) > 0 and datatype_name != "Select option":
                    if datatype_name == "String":
                        try:
                            uploaded_data[selected_column] = uploaded_data[selected_column].astype(str)
                        except ValueError as e:
                            # print(f"ValueError: {e}")
                            return jsonify(
                                {"status": "error", "message": f"Type conversion error for String type data!"})
                    elif datatype_name == "Integer":
                        try:
                            uploaded_data[selected_column] = uploaded_data[selected_column].astype(int)
                        except ValueError as e:
                            # print(f"ValueError: {e}")
                            return jsonify(
                                {"status": "error", "message": f"Type conversion error for Integer type data!"})
                    elif datatype_name == "Floating point":
                        try:
                            uploaded_data[selected_column] = uploaded_data[selected_column].astype(float)
                        except ValueError as e:
                            # print(f"ValueError: {e}")
                            return jsonify(
                                {"status": "error", "message": f"Type conversion error for Floating point type data!"})
                    elif datatype_name == "Datetime":
                        try:
                            uploaded_data[selected_column] = pd.to_datetime(uploaded_data[selected_column])
                        except ValueError as e:
                            # print(f"ValueError: {e}")
                            return jsonify(
                                {"status": "error", "message": f"Type conversion error for Datetime type data!"})
            elif method_name == "Date parsing":
                try:
                    uploaded_data[selected_column] = pd.to_datetime(uploaded_data[selected_column], errors='coerce')
                    # uploaded_data['date'] = [str(x) for x in list(uploaded_data[selected_column].dt.date)]
                    # uploaded_data['time'] = [str(x) for x in list(uploaded_data[selected_column].dt.time)]
                    if pd.api.types.is_datetime64_any_dtype(uploaded_data[selected_column]):
                        # Extract date and time
                        list1 = uploaded_data[selected_column].dt.date
                        list2 = uploaded_data[selected_column].dt.time
                        uploaded_data['Date'] = list1
                        uploaded_data['Time'] = list2
                    else:
                        # print("The column is not of datetime type. Please check the conversion.")
                        return jsonify({"status": "error", "message": f"Date parsing error!"})
                except Exception as e:
                    # print(f"Error: {e}")
                    return jsonify({"status": "error", "message": f"{e}!"})
        else:
            return jsonify({"status": "error", "message": f"Please select correct column!"})
        # return Response(status=204)
        # save the current data:
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}!"})

#handling state and city filter
@app.route('/handle_state_and_city', methods=["POST"])
@login_required
def handle_state_and_city():
    """
    state_column= request.form.get('dropdown7')
    cities_column= request.form.get('dropdown_city')
    """
    data = request.get_json()
    state_column = data.get('dropdown7')
    cities_column = data.get('dropdown_city')
    global uploaded_data
    global duplicate_data
    try:
        if (len(state_column) > 0 and state_column != "Select option" and state_column != "None") and (
                len(cities_column) > 0 and cities_column != "Select option" and cities_column != "None"):

            for x, y in zip(uploaded_data[state_column], uploaded_data[cities_column]):
                
                if x in state_and_city_dictionary.keys():
                    if y in state_and_city_dictionary[x]:
                        continue
                    else:
                        row_to_drop = uploaded_data[uploaded_data[cities_column] == y].index
                        duplicate_data = pd.concat([duplicate_data, uploaded_data.iloc[row_to_drop]], ignore_index=True)
                        uploaded_data = uploaded_data.drop(row_to_drop, axis=0).reset_index(drop=True)
                else:
                    row_to_drop = uploaded_data[uploaded_data[state_column] == x].index
                    duplicate_data = pd.concat([duplicate_data, uploaded_data.iloc[row_to_drop]], ignore_index=True)
                    uploaded_data = uploaded_data.drop(row_to_drop, axis=0).reset_index(drop=True)
            # print("uploaded_data:",uploaded_data,len(uploaded_data))
            # print("duplicate date:",duplicate_data,len(duplicate_data))
        else:
            return jsonify({"status": "error", "message": f"Please select correct column!"})
        # return Response(status=204)
        # save the current data:
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}!"})

#sending cities list to frontend
@app.route('/get_cities', methods=['POST'])
@login_required
def get_cities():
    statename = request.form.get('stateName')
    # print(statename)
    # print(state_and_city_dictionary.keys())
    if statename in list(state_and_city_dictionary.keys()):
        return jsonify({'cities': state_and_city_dictionary[statename]})
    else:
        return jsonify({'cities': []})

#removing rows or columns
@app.route('/remove_data', methods=["POST"])
@login_required
def remove_data():
    # handle column range if column not selected
    """
    selected_column = request.form.get('dropdown8')
    dropdown_format = request.form.get('dropdown_format')
    dropdown_type = request.form.get('dropdown_type')
    dropdown_remove_by = request.form.get('dropdown_remove_by')
    remove_input1 = request.form.get('remove_input1')
    remove_input2 = request.form.get('remove_input2')
    remove_input3 = request.form.get('remove_input3')
    """
    data = request.get_json()
    selected_column = data.get('dropdown8')
    dropdown_format = data.get('dropdown_format')
    dropdown_type = data.get('dropdown_type')
    dropdown_remove_by = data.get('dropdown_remove_by')
    remove_input1 = data.get('remove_input1')
    remove_input2 = data.get('remove_input2')
    remove_input3 = data.get('remove_input3')
    #treamming extra space in code
    remove_input1 = remove_input1.strip()
    remove_input2 = remove_input2.strip()
    remove_input3 = remove_input3.strip()
    global uploaded_data
    # print(selected_column,dropdown_format,dropdown_type,dropdown_remove_by,remove_input1,remove_input2,remove_input3)
    try:
        if selected_column != "Select option" and selected_column != "None" and dropdown_format != "Select option" and dropdown_type != "Select option" and dropdown_remove_by != "Select option":
            selected_column = selected_column.strip()
            if dropdown_format == "Row":
                if dropdown_type == "Date":
                    if dropdown_remove_by == "Value":
                        # Convert to datetime
                        uploaded_data[selected_column] = pd.to_datetime(uploaded_data[selected_column])  
                        #convert into time 2024-12-09 00:00:00
                        date_to_remove = pd.Timestamp(remove_input1)
                        #stores only not matched data
                        uploaded_data = uploaded_data[uploaded_data[selected_column] != date_to_remove]
                        #converting into string type
                        uploaded_data[selected_column] = uploaded_data[selected_column].astype(str)
                    elif dropdown_remove_by == "Range": 
                        uploaded_data[selected_column] = pd.to_datetime(uploaded_data[selected_column])
                        start_date = pd.Timestamp(remove_input2)
                        end_date = pd.Timestamp(remove_input3)
                        #skips the data greater than start_date and less than end_date
                        uploaded_data = uploaded_data[~((uploaded_data[selected_column] >= start_date) & (
                                    uploaded_data[selected_column] <= end_date))]
                        uploaded_data[selected_column] = uploaded_data[selected_column].astype(str)
                elif dropdown_type == "Number":
                    if dropdown_remove_by == "Value":
                        #stores not matched value data
                        uploaded_data = uploaded_data[uploaded_data[selected_column] != remove_input1]
                    elif dropdown_remove_by == "Range":
                        start_value = remove_input2
                        end_value = remove_input3
                        uploaded_data = uploaded_data[~((uploaded_data[selected_column] >= start_value) & (
                                    uploaded_data[selected_column] <= end_value))]
                elif dropdown_type == "Other":
                    if dropdown_remove_by == "Value":
                        uploaded_data = uploaded_data[uploaded_data[selected_column] != remove_input1]

            elif dropdown_format == "Column":
                uploaded_data = uploaded_data.drop(columns=[selected_column], axis=1)
        elif selected_column == "Select option" and dropdown_format == "Column" and dropdown_type == "Other":
            #used to remove column specified in the value
            if dropdown_remove_by == "Value":
                if remove_input1 != "":
                    if remove_input1 in uploaded_data.columns:
                        uploaded_data = uploaded_data.drop(columns=[remove_input1], axis=1)
                    else:
                        return jsonify(
                            {"status": "error", "message": f"Please enter correct column, column not found!"})
                else:
                    return jsonify({"status": "error", "message": f"Please enter correct column!"})
            elif dropdown_remove_by == "Range":
                if remove_input2 != "" and remove_input3 != "":
                    if remove_input2 in uploaded_data.columns and remove_input3 in uploaded_data.columns:
                        start_idx = uploaded_data.columns.get_loc(remove_input2)
                        end_idx = uploaded_data.columns.get_loc(remove_input3)
                        # print("start_idx",start_idx,"end_idx",end_idx)
                        if start_idx < end_idx:
                            cols_to_drop = uploaded_data.columns[start_idx:end_idx + 1].tolist()
                            uploaded_data = uploaded_data.drop(columns=cols_to_drop, axis=1)
                            # print("in start_idx", start_idx, "in end_idx", end_idx)
                        else:
                            return jsonify(
                                {"status": "error", "message": f"Please enter range like start index < end index !"})
                    else:
                        return jsonify({"status": "error", "message": f"Please enter correct columns!"})
                else:
                    return jsonify({"status": "error", "message": f"Please fill required fields!"})
        else:
            return jsonify({"status": "error", "message": f"Please select valid options!"})
        # return Response(status=204)
        # save the current data:
        uploaded_data.to_csv("./current_data/current_data.csv", index=False)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}!"})

# handle refresh button of data preperation
@app.route('/refreshData', methods=['POST'])
@login_required
def refreshData():
    # print("refreshData called")
    global uploaded_data
    if len(uploaded_data) == 0:
        if os.path.exists("./current_data/current_data.csv"):
            uploaded_data = pd.read_csv("./current_data/current_data.csv")
    return jsonify({"status": "success"})

"""
                                                End Of Data Preprocessing
                                                Start of Remove Duplicates
"""

#sending column  names as list 
@app.route('/removeDuplicate')
@login_required
def removeDuplicate():
    global global_uploaded_data
    global_uploaded_data = uploaded_data.copy()
    # print("in remove duplicate uploaded_data: ",len(uploaded_data))
    # print("in remove duplicate columns: ",uploaded_data.columns.tolist())

    if len(uploaded_data) > 0:
        column_headings = uploaded_data.columns.tolist()
        # print("in remove duplicate column_headings: ",column_headings)
        return render_template("components/remove_duplicate.html", column_headings=column_headings)
    else:
        column_headings = []
        return render_template("components/remove_duplicate.html", column_headings=column_headings)

#used to remove the duoplicate data after Remove Dupli button clicked
@app.route('/removeDuplicateData', methods=['POST'])
@login_required
def removeDuplicateData():
    data = request.get_json()
    selected_columns = data.get('textarea1')
    duplicate_filter_columns = data.get('textarea2')
    dropdown_operation = data.get('textarea3')

    try:
        if selected_columns and duplicate_filter_columns:
            task_id = f'task_{int(time.time())}'  # Unique task ID based on current timestamp
            tasks[task_id] = 'Processing'  # Set initial task status

            # Define the task in a separate thread
            def long_task():
                global uploaded_data
                global global_duplicated_data_all

                try:
                    # Specifies columns to keep in the final dataset 
                    selected_columns_list = [x.strip() for x in selected_columns.split(",")] 
                    # Specifies columns to check for duplicates
                    duplicate_filter_columns_list = [x.strip() for x in duplicate_filter_columns.split(",")]
                    # Operation type, either AND or OR.
                    dropdown_operation_list = dropdown_operation.split(",")

                    # Remove unwanted columns from uploaded data
                    data_fields = list(uploaded_data.columns)
                    remaining_column = set(data_fields).difference(set(selected_columns_list))
                    uploaded_data = uploaded_data.drop(remaining_column, axis=1)

                    # Database connection
                    if os.path.exists("./database_connection/database_connection.pkl"):
                        with open("./database_connection/database_connection.pkl", "rb") as f:
                            conn_data = pickle.load(f)
                    else:
                        conn_data = {}

                    client = MongoClient(
                        conn_data['database_connection'],
                        tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
                    coll_name_temp = conn_data['source_type'] + "_" + conn_data['collection_name']
                    coll_name_temp = coll_name_temp.lstrip("_")
                    collection = client[coll_name_temp]

                    # Remove leading/trailing whitespace in the selected columns
                    for col in duplicate_filter_columns_list:
                        uploaded_data[col] = uploaded_data[col].str.strip()

                    # Handle AND operation
                    if dropdown_operation_list[0] == "AND":
                        #identifying the duplicates matches all dupliate columns match first ocuurance kept last removed
                        global_duplicated_data_all = uploaded_data[uploaded_data.duplicated(subset=duplicate_filter_columns_list, keep='first')]
                        #remove duplicate rows matches duplicate columns match and lask ocuurance kept earlier removed
                        uploaded_data = uploaded_data.drop_duplicates(subset=duplicate_filter_columns_list, keep="last")

                        #changing the index values after dropping
                        uploaded_data = uploaded_data.reset_index(drop=True)

                        indices_to_remove = []
                        for i in range(len(uploaded_data)):
                            # accesing rows based on index
                            doc_data = uploaded_data.iloc[i]
                            #created dictionary with duplicates list column as key
                            conditions = [{column: str(doc_data[column])} for column in duplicate_filter_columns_list]
                            query = {"$and": conditions}
                            result = collection.find_one(query)
                            #checking already data present in database collection if it is present add it indices_remove_list
                            if result is not None:
                                indices_to_remove.append(i)
                        
                        # if duplicates is length is Zero and indices_to_remove  are there then we place it into duplicated data
                        if len(global_duplicated_data_all) == 0 and len(indices_to_remove) > 0:
                            global_duplicated_data_all = uploaded_data.iloc[indices_to_remove]
                            #reseting the indexs
                            global_duplicated_data_all.reset_index(drop=True, inplace=True)
                        else:
                            if len(indices_to_remove) > 0:
                                global_duplicated_data_all = pd.concat([global_duplicated_data_all, uploaded_data.iloc[indices_to_remove]])
                                #reset the indexs
                                global_duplicated_data_all.reset_index(drop=True, inplace=True)

                        uploaded_data.drop(indices_to_remove, inplace=True)
                        uploaded_data.reset_index(drop=True, inplace=True)

                    # Handle OR operation
                    elif dropdown_operation_list[0] == "OR":
                        temp_list = []
                        for col in duplicate_filter_columns_list:
                            temp_val = uploaded_data.duplicated(subset=[col], keep='first')
                            temp_list.append(temp_val)

                        mask_combined = temp_list[0]
                        for x1 in temp_list[1:]:
                            mask_combined = mask_combined | x1

                        global_duplicated_data_all = uploaded_data[mask_combined]
                        uploaded_data = uploaded_data[~mask_combined]
                        uploaded_data = uploaded_data.reset_index(drop=True)

                        indices_to_remove = []
                        for i in range(len(uploaded_data)):
                            doc_data = uploaded_data.iloc[i]
                            conditions = [{column: str(doc_data[column])} for column in duplicate_filter_columns_list]
                            query = {"$or": conditions}
                            result = collection.find_one(query)
                            if result is not None:
                                indices_to_remove.append(i)

                        if len(global_duplicated_data_all) == 0 and len(indices_to_remove) > 0:
                            global_duplicated_data_all = uploaded_data.iloc[indices_to_remove]
                            global_duplicated_data_all.reset_index(drop=True, inplace=True)
                        else:
                            if len(indices_to_remove) > 0:
                                global_duplicated_data_all = pd.concat([global_duplicated_data_all, uploaded_data.iloc[indices_to_remove]])
                                global_duplicated_data_all.reset_index(drop=True, inplace=True)

                        uploaded_data.drop(indices_to_remove, inplace=True)
                        uploaded_data.reset_index(drop=True, inplace=True)

                    else:
                        tasks[task_id] = 'Error'
                        return jsonify({'status': 'error', 'message': "Please select the correct method!"})

                    # Save unique data
                    folder_path = "./unique_files"
                    if not os.path.exists(folder_path):
                        os.makedirs(folder_path)
                    now = datetime.datetime.now()
                    date_str = now.strftime("%Y%m%d")
                    new_filename = f"{conn_data['source_type']}_{conn_data['collection_name']}_unique_file_{date_str}.csv"
                    new_filename = new_filename.lstrip("_")
                    file_path = os.path.join(folder_path, new_filename)
                    if len(uploaded_data) > 0:
                        uploaded_data.to_csv(file_path, index=False)

                    # Save duplicate data
                    folder_path_dup = "./duplicate_files"
                    if not os.path.exists(folder_path_dup):
                        os.makedirs(folder_path_dup)
                    new_filename_dup = f"{conn_data['source_type']}_{conn_data['collection_name']}_duplicate_file_{date_str}.csv"
                    new_filename = new_filename.lstrip("_")
                    file_path_dup = os.path.join(folder_path_dup, new_filename_dup)
                    if len(global_duplicated_data_all) > 0:
                        global_duplicated_data_all.to_csv(file_path_dup, index=False)

                    # Save current data
                    if len(uploaded_data) > 0:
                        uploaded_data.to_csv("./current_data/current_data.csv", index=False)

                    tasks[task_id] = 'Completed'
                except Exception as e:
                    tasks[task_id] = f'Error: {str(e)}'

            # Start the task asynchronously
            Thread(target=long_task).start()

            return jsonify({"status": "success", "task_id": task_id})
        else:
            return jsonify({"status": "error", "message": "Please select correct columns!"})
    except Exception as e:
        error_message = str(e)
        return jsonify({"status": "error", "message": f"Error occurred: {e}!"})

#sending count of unique data and duplicated data shown in model
@app.route('/check_task_status_remove_duplicate/<task_id>', methods=['GET'])
@login_required
def check_task_status_remove_duplicate(task_id):
    task_status = tasks.get(task_id, 'Task not found')
    if task_status == 'Completed':
        unique_data_count = len(uploaded_data)
        duplicate_data_count = len(global_duplicated_data_all)
        return jsonify({'status': 'Completed', 'unique_data': unique_data_count, 'duplicate_data': duplicate_data_count})
    return jsonify({'status': task_status})

#showing the data after show button clicked
@app.route('/removeShowdata')
@login_required
def removeShowdata():
    global uploaded_data
    # Check if the data is empty
    if uploaded_data.empty:
        return jsonify({'len_is_zero': True})  # Return special message if no data
    else:
        data_table = uploaded_data.to_html(classes='dataframe', index=False, escape=False)
        return jsonify({'data_table': data_table, 'len_is_zero': False})


#saving data in database after clicking save in DB button
@app.route('/save_in_DB', methods=['POST'])
@login_required
def save_in_DB():
    global uploaded_data
    try:
        if len(uploaded_data) == 0:
            tasks["task_1"] = "Error: No data to insert into Database!"
            return jsonify({"status": "error", "message": "No data to insert into Database!"}), 500

        task_id = f'task_{int(time.time())}'  # Unique task ID based on current timestamp
        tasks[task_id] = 'Processing'

        def long_task():
            try:
                if os.path.exists("./database_connection/database_connection.pkl"):
                    directory = './database_connection'
                    file_path = os.path.join(directory, 'database_connection.pkl')
                    with open(file_path, "rb") as f:
                        conn_data = pickle.load(f)

                    client = MongoClient(
                        conn_data['database_connection'],
                        tlsCAFile=certifi.where(), tls=True
                    ).get_database(conn_data['database_name'])

                    coll_name_temp = conn_data['source_type'] + "_" + conn_data['collection_name']
                    coll_name_temp = coll_name_temp.lstrip("_")
                    collection = client[coll_name_temp]

                    if 'phone' not in uploaded_data.columns.tolist():
                        tasks[task_id] = "Error: 'phone' column doesn't exist!"
                        return

                    phones = uploaded_data['phone'].tolist()
                    existing_phones = collection.find({"phone": {"$in": phones}}, {"phone": 1})
                    existing_phones_set = {doc['phone'] for doc in existing_phones}
                    rows_to_keep = uploaded_data[~uploaded_data['phone'].isin(existing_phones_set)]

                    if rows_to_keep.empty:
                        tasks[task_id] = "Error: No new data to insert into Database!"
                        return

                    pipeline = [{'$group': {'_id': None, 'max_value': {'$max': '$id'}}}]
                    result = list(collection.aggregate(pipeline))
                    max_value = result[0]['max_value'] if result else 0

                    id_list = list(range(max_value + 1, max_value + 1 + len(rows_to_keep)))
                    if 'id' not in rows_to_keep.columns:
                        rows_to_keep.insert(0, 'id', id_list)
                    else:
                        rows_to_keep['id'] = id_list

                    data_for_insert = rows_to_keep.to_dict(orient='records')
                    insert_result = collection.insert_many(data_for_insert)

                    if insert_result.acknowledged:
                        tasks[task_id] = "Completed"
                    else:
                        tasks[task_id] = "Error: Data insertion failed!"

                else:
                    tasks[task_id] = "Error: Database connection file not found!"

            except Exception as e:
                tasks[task_id] = f"Error: {str(e)}"

        Thread(target=long_task).start()

        return jsonify({"status": "success", "task_id": task_id})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/task_status_save_DB/<task_id>', methods=['GET'])
@login_required
def task_status(task_id):
    try:
        if task_id in tasks:
            task_status = tasks[task_id]
            if task_status == 'Completed':
                return jsonify({"status": "Completed"})
            elif task_status == 'Processing':
                return jsonify({"status": "Processing"})
            elif "Error" in task_status:
                return jsonify({"status": "Error", "message": task_status})
            else:
                return jsonify({"status": "Unknown", "message": "Unknown task status."})
        else:
            return jsonify({"status": "error", "message": "Task not found."}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


""" 
                                                    End remove duplicate 
                                                    End of Deduplication
                                                    Start Export Data
"""

# start data exporting
@app.route('/exportData')
@login_required
def exportData():
    return render_template("components/export_data.html")

#sending unique and duplicated files based on selecting type in Export page
@app.route('/get_files', methods=['POST'])
@login_required
def get_files():
    unique_files = []
    duplicate_files = []
    if os.path.exists("./unique_files"):
        unique_files = os.listdir("./unique_files")
    if os.path.exists("./duplicate_files"):
        duplicate_files = os.listdir("./duplicate_files")
    file_type = request.form.get('file_type')
    if file_type == 'Unique data':
        return jsonify({'files': unique_files}) 
    elif file_type == 'Duplicate data':
        return jsonify({'files': duplicate_files})
    else:
        return jsonify({'files': []})

#download the selected file

@app.route('/export_download', methods=['POST'])
@login_required
def export_download():
    file_type = request.form.get('file_type')
    file_name = request.form.get('file_name')
    # print("file_type:",file_type,"file_name:",file_name)
    if (len(file_type) > 0 and file_type != "Select option") and (len(file_name) > 0 and file_name != "Select option"):
        if file_type == "Unique data":
            file_path = os.path.join("./unique_files", file_name)
            try:
                return send_file(file_path, as_attachment=True)
            except FileNotFoundError:
                abort(404)  # Return 404 if file not found
        elif file_type == "Duplicate data":
            file_path = os.path.join("./duplicate_files", file_name)
            try:
                return send_file(file_path, as_attachment=True)
            except FileNotFoundError:
                abort(404)  # Return 404 if file not found
    else:
        return Response(status=204)

#showing the selected file data
@app.route('/datashow', methods=['POST'])
@login_required
def datashow():
    # print("method call for export data show")
    file_type = request.form.get('file_type')
    file_name = request.form.get('file_name')
    if file_type == 'Unique data':
        file_path = os.path.join("./unique_files", file_name)
        # print("file_path:",file_path)
        try:
            df = pd.read_csv(file_path)
            data_table_html = df.to_html(classes='display', index=False)
            return jsonify(data_table=data_table_html)
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
            data_table_html = df.to_html(classes='display', index=False)
            return jsonify(data_table=data_table_html)
    elif file_type == 'Duplicate data':
        file_path = os.path.join("./duplicate_files", file_name)
        try:
            df = pd.read_csv(file_path)
            data_table_html = df.to_html(classes='display', index=False)
            return jsonify(data_table=data_table_html)
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
            data_table_html = df.to_html(classes='display', index=False)
            return jsonify(data_table=data_table_html)

#remove the selected file
@app.route('/deletefile', methods=['POST'])
@login_required
def deletefile():
    file_type = request.form.get('file_type')
    file_name = request.form.get('file_name')
    if file_type == 'Unique data':
        file_path = os.path.join("./unique_files", file_name)
        os.remove(file_path)
    elif file_type == 'Duplicate data':
        file_path = os.path.join("./duplicate_files", file_name)
        os.remove(file_path)
    success_message = f"File {file_name} of type {file_type} deleted successfully."
    return jsonify(message=success_message)

"""
                                                    Filter records base on column
"""
#If Edit Button clicked show some form code wriiten in js 
#when i Clicked Refresh button in the form this navigates to this url....

@app.route('/getItems', methods=['POST'])
def getItems():
    # start remove invalid file 
    older_path = "./ExportDataEditFile"
    file_path = "wrong_data_exportDataEditFile.csv"
    complete_path = os.path.join(older_path, file_path)
    if len(ExportDataEditOkNew) > 0:  # making empty the dataframe
        ExportDataEditOkNew.drop(ExportDataEditOkNew.index, inplace=True)

    #remove the path if already present
    if os.path.exists(complete_path):
        os.remove(complete_path)
    # end remove invalid file
    global uploaded_data
    global ExportDataEditOkOriginal
    status = request.form.get('status')
    if status == "true":
        try:
            file_type = request.form.get('file_type')
            file_name = request.form.get('file_name')
            df = file_reading(file_type, file_name)
            ExportDataEditOkOriginal = df
            items = df.columns.tolist()
            return jsonify({"status": "success", "items": items})
        except Exception as e:
            return jsonify({"status": "error", "message": f"{e}"}), 500
    else:
        if len(ExportDataEditOkOriginal) > 0:
            items = ExportDataEditOkOriginal.columns.tolist()
            return jsonify({"status": "success", "items": items})
        elif len(uploaded_data) > 0:
            items = uploaded_data.columns.tolist()
            return jsonify({"status": "success", "items": items})
        else:
            return jsonify({"status": "error", "message": "There is no data"}), 500
        # return jsonify(items)

#checking the file where it present
def file_reading(file_type, file_name):
    if file_type == 'Unique data':
        file_path = os.path.join("./unique_files", file_name)
        try:
            df = pd.read_csv(file_path)
            return df
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
            return df
    elif file_type == 'Duplicate data':
        file_path = os.path.join("./duplicate_files", file_name)
        try:
            df = pd.read_csv(file_path)
            return df
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
            return df

#handling the Export Edit Ok button 
@app.route('/exportEditOk', methods=['POST'])
@login_required
def exportEditOk():
    file_type = request.form.get('file_type')
    file_name = request.form.get('file_name')
    checkbox0 = request.form.get('checkbox0')
    checkbox1 = request.form.get('checkbox1')
    checkbox2 = request.form.get('checkbox2')
    checkbox3 = request.form.get('checkbox3')
    checkbox4 = request.form.get('checkbox4')
    checkbox5 = request.form.get('checkbox5')
    checkbox6 = request.form.get('checkbox6')
    checkbox7 = request.form.get('checkbox7')
    checkbox8 = request.form.get('checkbox8')
    checkbox9 = request.form.get('checkbox9')
    checkbox10 = request.form.get('checkbox10')
    column_name = request.form.get('input_field_column_name_handle')
    input_field_customTextarea = request.form.get('input_field_customTextarea')
    input_field_input61 = request.form.get('input_field_input61')
    input_field_input71 = request.form.get('input_field_input71')
    input_field_input72 = request.form.get('input_field_input72')
    input_field_input81 = request.form.get('input_field_input81')
    input_field_input82 = request.form.get('input_field_input82')
    input_field_input91 = request.form.get('input_field_input91')
    input_field_input92 = request.form.get('input_field_input92')
    input_field_input101 = request.form.get('input_field_input101')
    input_field_input102 = request.form.get('input_field_input102')
    checkbox_list = [checkbox1, checkbox2, checkbox3, checkbox4, checkbox5, checkbox6, checkbox7, checkbox8, checkbox9,
                     checkbox10]
    all_bool_data = tuple(True if x == "true" else False for x in checkbox_list)
    checkbox1, checkbox2, checkbox3, checkbox4, checkbox5, checkbox6, checkbox7, checkbox8, checkbox9, checkbox10 = all_bool_data
    if sum(all_bool_data) == 0:
        return jsonify({"status": "error", "message": f"Please select operations to perform!"})
    # performing operations
    """
    data = {
        'Name': ['Aliceee', 'Bob', 'Chaaaarlie', 'Davaaid', 'Eva','sdad','APP'],
        'Age': [23, 35, 45, 22, 30,55,60],
        'City': ['New York45 aa dfdf dfdfdf ppp', 'Los Angeles45545 sds sd', 'pa234pi ss', 'Houston aa vv', 'Phoenix','hi','hello']
    }

    df = pd.DataFrame(data)
    """
    global uploaded_data
    global ExportDataEditOkOriginal
    global ExportDataEditOkNew
    global ExportDataEditUploadFileData
    #if already data present removing first
    if len(ExportDataEditUploadFileData)>0:
        ExportDataEditUploadFileData = ExportDataEditUploadFileData.drop(ExportDataEditUploadFileData.index).drop(ExportDataEditUploadFileData.columns, axis=1)
        
    # print("column name: ",column_name)
    if column_name != "None" and column_name != "":
        if checkbox0 == "true":
            if file_type == "Unique data":
                file_path = f"./unique_files/{file_name}"
                try:
                    ExportDataEditOkOriginal = pd.read_csv(file_path)
                    print(ExportDataEditOkOriginal.shape)
                    ExportDataEditOkNew = pd.DataFrame(columns=ExportDataEditOkOriginal.columns.tolist())
                except Exception as e:
                    return jsonify({"status": "error", "message": f"{e}"})
            else:
                return jsonify({"status": "error", "message": f"Please select the unique file!"})
        else:
            if len(uploaded_data) > 0:
                ExportDataEditOkOriginal = uploaded_data
                ExportDataEditOkNew = pd.DataFrame(columns=uploaded_data.columns.tolist())
            else:
                return jsonify({"status": "error", "message": f"Empty dataframe!"})

        # executiong of code
        if len(ExportDataEditOkOriginal) > 0:
            if checkbox1:
                masking = ExportDataEditOkOriginal[column_name].apply(vowelChecking)
                # stores only masking value is true
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox2:
                masking = ExportDataEditOkOriginal[column_name].apply(palindromeChecking)
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox3:
                masking = ExportDataEditOkOriginal[column_name].apply(whiteSpaceChecking)
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox4:
                masking = ExportDataEditOkOriginal[column_name].apply(caseSensitiveChecking)
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox5:
                input_field_customTextarea = [x.strip() for x in input_field_customTextarea.split(',')]
                masking = ExportDataEditOkOriginal[column_name].apply(customValueChecking,
                                                                      list_of_values=input_field_customTextarea)
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox6:
                digit_limit = int(input_field_input61)
                masking = ExportDataEditOkOriginal[column_name].apply(numericalChecking, digit_limit=digit_limit)
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox7:
                masking = ExportDataEditOkOriginal[column_name].apply(specialCharsChecking,
                                                                      from_special_char_count=int(input_field_input71),
                                                                      to_special_char_count=int(input_field_input72))
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox8:
                masking = ExportDataEditOkOriginal[column_name].apply(repeatedCharsChecking,
                                                                      min_repeats=int(input_field_input81),
                                                                      max_repeats=int(input_field_input82))
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]

            if checkbox9:
                masking = ExportDataEditOkOriginal[column_name].apply(charLengthChecking,
                                                                      min_char_count=int(input_field_input91),
                                                                      max_char_count=int(input_field_input92))
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
            if checkbox10:
                #print("testing option")
                #print("data length: ",len(ExportDataEditOkOriginal))

                masking = ExportDataEditOkOriginal[column_name].apply(wordLengthChecking,
                                                                      min_word_count=int(input_field_input101),
                                                                      max_word_count=int(input_field_input102))
                # ExportDataEditOkNew.append(ExportDataEditOkOriginal.loc[masking])
                #print("testing option2")
                ExportDataEditOkNew = pd.concat([ExportDataEditOkNew, ExportDataEditOkOriginal.loc[masking]],
                                                ignore_index=True)
                ExportDataEditOkOriginal = ExportDataEditOkOriginal[~masking]
                #print("ExportDataEditOkNew: ",ExportDataEditOkNew)
                #print("ExportDataEditOkOriginal: ", ExportDataEditOkOriginal)
            # saving the data in file
            folder_path = "./ExportDataEditFile"
            #already exists remove it
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path)
            #if not exists create it
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            file_path = os.path.join(folder_path, "wrong_data_exportDataEditFile.csv")
            if len(ExportDataEditOkNew) > 0:
                ExportDataEditOkNew.to_csv(file_path, index=False)
            else:
                return jsonify({"status": "error", "message": f"No records selected for your conditions!"})
        else:
            return jsonify({"status": "error", "message": f"Empty Dataframe!"})

    else:
        return jsonify({"status": "error", "message": f"Please select column for performing operation!"})

    return jsonify({"status": "success"})

# cut out records based on pattern matching

#checking value not contains single vowel
def vowelChecking(column_value):
    #print("vowel checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        vowels_count = len(re.findall(r'[AEIOUaeiou]', column_value))
        if vowels_count == 0:
            return True
        else:
            return False
    except Exception as e:
        return False

#checking palindrome rows
def palindromeChecking(column_value):
    #print("palindrome checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        if column_value == column_value[::-1]:
            return True
        else:
            return False
    except Exception as e:
        return False

#checking leading or trailing whitespaces
def whiteSpaceChecking(column_value):
    #print("WhiteSpace checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None:  # NaN check
            column_value = ""
        if column_value != column_value.strip() or "  " in column_value:  # Leading/trailing or double spaces
            return True
        else:
            return False
    except Exception as e:
        return False

#checking case either all upper or all lovwer case letters
def caseSensitiveChecking(column_value):
    #print("case sensitive checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        if column_value.isupper() or column_value.islower():
            return True
        else:
            return False
    except Exception as e:
        return False

#checking value entered by user for checking..
def customValueChecking(column_value, list_of_values):
    #print("custom Value checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        if column_value in list_of_values:
            return True
        else:
            return False
    except Exception as e:
        return False

#checking count of nummerical in the value return true if count greater than specified count
def numericalChecking(column_value, digit_limit):
    #print("Numerical checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        if len(re.findall(r'\d', column_value)) >= digit_limit:
            return True
        else:
            return False
    except Exception as e:
        return False

#using regex compare value
def specialCharsChecking(column_value, from_special_char_count, to_special_char_count):
    #print("special chars checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        #Regex for special characters it finds the length not having A_Z or a-z and whitespaces
        special_chars_count = len(re.findall(r'[^A-Za-z\s]', column_value))
        if from_special_char_count > 0 and to_special_char_count > 0:
            if special_chars_count >= from_special_char_count and special_chars_count <= to_special_char_count:
                return True
            else:
                return False
        elif from_special_char_count > 0 and to_special_char_count == 0:
            if special_chars_count >= from_special_char_count:
                return True
            else:
                return False
        elif from_special_char_count == 0 and to_special_char_count > 0:
            if special_chars_count <= to_special_char_count:
                return True
            else:
                return False
        else:
            return False
    except Exception as e:
        return False

#using regex compare value
def repeatedCharsChecking(column_value, min_repeats, max_repeats):
    #print("repeated chars checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        if min_repeats > 0 and max_repeats > 0:
            pattern = r'(.)\1{' + str(min_repeats - 1) + ',' + str(max_repeats - 1) + '}'
            if re.search(pattern, column_value):
                return True
            else:
                return False
        elif min_repeats > 0 and max_repeats == 0:
            pattern = r'(.)\1{' + str(min_repeats - 1) + ',}'
            if re.search(pattern, column_value):
                return True
            else:
                return False
        elif min_repeats == 0 and max_repeats > 0:
            pattern = r'(.)\1{' + str(max_repeats - 1) + '}'
            if re.search(pattern, column_value):
                return True
            else:
                return False
        return False
    except Exception as e:
        return False

#using len() method compare
def charLengthChecking(column_value, min_char_count, max_char_count):
    #print("Char length checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        if min_char_count > 0 and max_char_count > 0:
            if len(column_value) >= min_char_count and len(column_value) <= max_char_count:
                return True
            else:
                return False
        elif min_char_count > 0 and max_char_count == 0:
            if len(column_value) >= min_char_count:
                return True
            else:
                return False
        elif min_char_count == 0 and max_char_count > 0:
            if len(column_value) <= max_char_count:
                return True
            else:
                return False
        else:
            return False
    except Exception as e:
        return False

#split the value then find lenght
def wordLengthChecking(column_value, min_word_count, max_word_count):
    #print("word length checking called")
    try:
        if isinstance(column_value, float):
            column_value = str(column_value)
        if column_value is None or column_value != column_value:  # NaN check
            column_value = ""
        word_Count = column_value.split()
        if min_word_count > 0 and max_word_count > 0:
            if len(word_Count) >= min_word_count and len(word_Count) <= max_word_count:
                return True
            else:
                return False
        elif min_word_count > 0 and max_word_count == 0:
            if len(word_Count) >= min_word_count:
                return True
            else:
                return False
        elif min_word_count == 0 and max_word_count > 0:
            if len(word_Count) <= max_word_count:
                return True
            else:
                return False
        else:
            return False
    except Exception as e:
        return False

# download the filtered data
@app.route('/editDownloadShow')
@login_required
def editDownloadShow():
    folder_path = "./ExportDataEditFile"
    file_path = "wrong_data_exportDataEditFile.csv"
    complete_path = os.path.join(folder_path, file_path)
    # print('complete_path: ',complete_path)
    if os.path.exists(complete_path):
        try:
            return send_file(complete_path, as_attachment=True)
        except Exception as e:
            # return jsonify({"status": "error", "message": str(e)}), 500
            pass
    else:
        # return jsonify({"status": "error", "message": "No file to download!"}), 500
        pass
    return Response(status=204)

#showing the invalid (filtered data) stored in ExportDataEditOkNew | wrong_data_exportDataEditFile.csv
@app.route('/invalidEditShow', methods=['POST'])
@login_required
def invalidEditShow():
    # status = request.form.get('status')
    global uploaded_data
    folder_path = "./ExportDataEditFile"
    file_path = "wrong_data_exportDataEditFile.csv"
    complete_path = os.path.join(folder_path, file_path)
    try:
        if os.path.exists(complete_path):
            # print("invalid 1")
            df = pd.read_csv(complete_path)
            data_table_html = df.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({"status": "success", "data_table": data_table_html})
        else:
            if len(ExportDataEditOkNew) > 0:
                # print("invalid 2")
                data_table_html = ExportDataEditOkNew.to_html(classes='display', index=False)
                # return jsonify(data_table=data_table_html)
                return jsonify({"status": "success", "data_table": data_table_html})
            else:
                # print("invalid 3")
                return jsonify(
                    {"status": "error", "message": f"Its data issue, please check data once, nothing to show!"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"})

#showing the data after filtering (valid data) stored in ExportDataEditOkOriginal | currentdata.csv | uploaded_data
@app.route('/validEditShow', methods=['POST'])
@login_required
def validEditShow():
    # print("ExportDataEditOkOriginal: ",ExportDataEditOkOriginal)
    # print("uploaded_data: ",uploaded_data)
    file_path = "./current_data/current_data.csv"
    try:
        if len(ExportDataEditOkOriginal) > 0:
            # print("valid 1")
            data_table_html = ExportDataEditOkOriginal.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({"status": "success", "data_table": data_table_html})
        elif len(ExportDataEditOkOriginal) == 0 and len(ExportDataEditOkNew)>0:
            print("part 1",ExportDataEditOkOriginal)
            data_table_html = ExportDataEditOkOriginal.to_html(classes='display', index=False)
            return jsonify({"status": "success", "data_table": data_table_html})
        elif os.path.exists(file_path):
            dataset = pd.read_csv(file_path)
            data_table_html = dataset.to_html(classes='display', index=False)
            if len(dataset) > 0:
                return jsonify({"status": "success", "data_table": data_table_html})
            """
            else:
                return jsonify({"status": "error", "message": "Empty dataframe!"})
            """
        else:
            if len(uploaded_data) > 0:
                # print("valid 2")
                data_table_html = uploaded_data.to_html(classes='display', index=False)
                # return jsonify(data_table=data_table_html)
                return jsonify({"status": "success", "data_table": data_table_html})
            else:
                # print("valid 3")
                return jsonify(
                    {"status": "error", "message": f"Its data issue, please check data once, nothing to show!"})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}, Its data issue, please check data once!"})
    return jsonify({"status": "error", "message": f"Its data issue, please check data once!"})

# upload file
@app.route('/editUploadFile', methods=['POST'])
@login_required
def editUploadFile():
    if 'file_input' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 500

    file = request.files['file_input']
    data_separator_symbol = request.form.get('text_input', ',')

    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'}), 500

    try:
        original_filename = file.filename
        base, extension = os.path.splitext(original_filename)
        data_separator_symbol = None if (data_separator_symbol == None or data_separator_symbol == '') and extension == ".csv" else data_separator_symbol

        folder_path = "./ExportDataEditUploadFile"
        #remove the invalid data 
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        new_file_path = os.path.join(folder_path, file.filename)


        global ExportDataEditUploadFileData
        if extension == ".xlsx":
            file.save(new_file_path)
            ExportDataEditUploadFileData = pd.read_excel(new_file_path)

        elif extension == ".csv":
            file.save(new_file_path)
            file_decode = detect_encoding(new_file_path)
            try:
                ExportDataEditUploadFileData = pd.read_csv(new_file_path, encoding=file_decode, sep=None,engine="python")
            except Exception as e:
                try:
                    ExportDataEditUploadFileData = pd.read_csv(new_file_path, sep=None, engine="python")
                except Exception as e1:
                    if file_decode not in list(set(encodings.aliases.aliases.values())):
                        return jsonify({'status': 'error', 'message': 'Encoding format not found'}), 500
                    else:
                        return jsonify({'status': 'error', 'message': f'{e1}'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'{e}'}), 500
    # print("file name: ",file.filename)
    # return Response(status=204)
    return jsonify({'status': 'success'})

@app.route('/uploadShowUploadedFile', methods=['POST', 'GET'])
@login_required
def uploadShowUploadedFile():
    # print("showUploadedFile")
    global ExportDataEditUploadFileData
    try:
        if len(ExportDataEditUploadFileData) > 0:
            data_table_html = ExportDataEditUploadFileData.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({'status': 'success', 'data_table': data_table_html})
        else:
            """
            folder_name="./ExportDataEditUploadFile"
            file_name=os.listdir(folder_name)

            if len(file_name)>0 and os.path.exists(folder_name):
                file_path=os.path.join(folder_name,file_name[0])
                data=pd.read_csv(file_path)
                ExportDataEditUploadFileData=data
                data_table_html = data.to_html(classes='display', index=False)
                return jsonify(data_table=data_table_html)
            else:
                return jsonify({'error': "No file exists"}), 500
            """
            # return jsonify({'error': "No file exists"}), 500
            return jsonify({'status': 'error', 'message': "Something problem with file/file data!"}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': f"{e}"}), 500

#merge the upload file to unique file 
@app.route('/mergeUploadedFileData', methods=['POST'])
@login_required
def mergeUploadedFileData():
    global ExportDataEditOkOriginal #unique data
    global ExportDataEditUploadFileData#new file uploaded data
    global uploaded_data #previous uploaded data at fileupload
    """
    file_type = request.form.get('file_type')
    file_name = request.form.get('file_name')
    print("file_type: ",file_type,"file_name: ",file_name)
    """
    print("Merge method called!")
    print(ExportDataEditOkOriginal.shape)
    if len(ExportDataEditOkOriginal) > 0 and len(ExportDataEditUploadFileData) > 0:
        if list(ExportDataEditOkOriginal.columns) == list(ExportDataEditUploadFileData.columns):
            #new code added
            key_columns = ExportDataEditOkOriginal.columns.tolist()
            mask = ~ExportDataEditUploadFileData.set_index(key_columns).index.isin(ExportDataEditOkOriginal.set_index(key_columns).index)
            df_to_append = ExportDataEditUploadFileData[mask]
            if len(df_to_append)==0:
                return jsonify({'status': 'error', 'message': "No data to merge, please check!"}), 500
            #end new code added
            """
            ExportDataEditOkOriginal = pd.concat([ExportDataEditOkOriginal, ExportDataEditUploadFileData],
                                                 ignore_index=True)
            """
            if len(df_to_append)>0:
                ExportDataEditOkOriginal = pd.concat([ExportDataEditOkOriginal, df_to_append],
                                                 ignore_index=True)
            return jsonify({'status': 'success'})
        else:
            return jsonify({'status': 'error', 'message': "Data format not matched!"}), 500
    elif len(ExportDataEditOkOriginal) == 0 and len(ExportDataEditUploadFileData) > 0:
            if len(ExportDataEditUploadFileData) > 0:
                ExportDataEditOkOriginal = ExportDataEditUploadFileData
                return jsonify({'status': 'success'})
            else:
                return jsonify({'status': 'error', 'message': "No data to merge, please check!"}), 500
    else:
        return jsonify({'status': 'error', 'message': "something wrong with data to merge, please check!"}), 500

#show the data after merging...
@app.route('/mergedShowUploadedFile', methods=['POST'])
@login_required
def mergedShowUploadedFile():
    global ExportDataEditOkOriginal
    try:
        if len(ExportDataEditOkOriginal) > 0:
            data_table_html = ExportDataEditOkOriginal.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({'status': 'success', 'data_table': data_table_html})
        else:
            # return jsonify({'error': "No file exists"}), 500
            return jsonify({'status': 'error', 'message': "Data not exists!"}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': f"{e}"}), 500

@app.route('/updateFileData', methods=['POST'])
@login_required
def updateFileData():
    global ExportDataEditOkOriginal
    global uploaded_data
    data = request.json
    file_type = data.get('file_type')
    file_name = data.get('file_name')
    # print("file_type: ",file_type,"file_name: ",file_name)
    try:
        if len(ExportDataEditOkOriginal) > 0:
            uploaded_data = ExportDataEditOkOriginal
            if os.path.exists("./current_data/current_data.csv"):
                ExportDataEditOkOriginal.to_csv("./current_data/current_data.csv", index=False)
            # update file
            if file_type == "Unique data":
                file_path = f"./unique_files/{file_name}"
                if os.path.exists(file_path):
                    data_read = pd.read_csv(file_path)
                    if len(data_read) == len(ExportDataEditOkOriginal):
                        ExportDataEditOkOriginal.to_csv(file_path, index=False)
                        return jsonify({'status': 'success'})
                    else:
                        return jsonify({'status': 'error', 'message': "Data length not matched!"}), 500
                else:
                    return jsonify({'status': 'error', 'message': "File not exists!"}), 500
            else:
                return jsonify({'status': 'error', 'message': "Please select unique file!"}), 500
        else:
            return jsonify(
                {'status': 'error', 'message': "Something wrong with data,Empty dataset, please check!"}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': f"{e}!"}), 500

"""
                                                Filter records base on column Ended
                                                Export Data Ends
                                                Starts the uploaded File
"""

# start handle uploaded file
@app.route('/uploadedFile')
@login_required
def uploadedFile():
    dst_folder = "./file_store"
    if os.path.exists(dst_folder):
        files = os.listdir(dst_folder)
        return render_template("components/uploaded_file.html", files=files)
    else:
        return render_template("components/uploaded_file.html", files=[])


@app.route('/uploaded_download_file/<filename>')
@login_required
def uploaded_download_file(filename):
    try:
        return send_from_directory(app.config['FILE_STORE'], filename, as_attachment=True)
    except FileNotFoundError:
        abort(404)


@app.route('/uploaded_delete_file/<filename>')
@login_required
def uploaded_delete_file(filename):
    try:
        os.remove(os.path.join(app.config['FILE_STORE'], filename))
        return redirect(url_for('uploadedFile'))
    except FileNotFoundError:
        abort(404)

"""
                                                Ends the uploaded File
                                                Truncate DB Starts
"""

#handling TruncateDB
@app.route('/truncate_database')
@login_required
def truncate_database():
    return render_template('components/truncate_database.html')

#showing data at Truncate DB
@app.route('/showTruncateDatabase', methods=['POST'])
@login_required
def showTruncateDatabase():
    # button_id = request.args.get('id')
    # test case
    try:
        """
        p=True
        if p:
            raise FileNotFoundError("database_connection.pkl file not found.")
        """
        # reading database connection
        if os.path.exists("./database_connection/database_connection.pkl"):
            directory = './database_connection'
            file_path = os.path.join(directory, 'database_connection.pkl')
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    conn_data = pickle.load(f)
            else:
                conn_data = {}

        client = MongoClient(
            conn_data['database_connection'],
            tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
        # collection_names = client[conn_data['collection_name']]
        coll_name_temp = conn_data['source_type'] + "_" + conn_data['collection_name']
        coll_name_temp = coll_name_temp.lstrip("_")
        collection_names = client[coll_name_temp]
        data = request.json
        file_type = data.get('file_type')
        file_name = data.get('file_name')
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        button_id = data.get('id')
        date_column = data.get('date_column')
        # add new
        from_id = int(data.get('from_id'))
        to_id = int(data.get('to_id'))
        id_column = data.get('id_column')
        # print("BUTTON REQUEST ID:",button_id)
        # print("inside show method")
        # print(file_type, file_name)
        #print(from_date, to_date, date_column)
        if button_id == "1":

            if file_type == "Unique data":
                file_path = os.path.join("./unique_files", file_name)
                data = pd.read_csv(file_path)
                data_table = data.to_html(classes='dataframe', index=False, escape=False)
                return jsonify({"status": "success", 'data_table': data_table})
            elif file_type == "Duplicate data":
                file_path = os.path.join("./duplicate_files", file_name)
                data = pd.read_csv(file_path)
                data_table = data.to_html(classes='dataframe', index=False, escape=False)
                return jsonify({"status": "success", 'data_table': data_table})
        elif button_id == "2":
            if (from_date != "" and to_date != "" and date_column != ""):
                date_column = date_column.strip(" ")
                date_column = date_column.strip("\n")
                results = collection_names.find({
                    date_column: {
                        '$gte': from_date,
                        '$lte': to_date
                    }
                }, {'_id': 0})
                data = list(results)
                if len(data) > 0:
                    df2 = pd.DataFrame(data)
                    print(df2)
                    data_table = df2.to_html(classes='dataframe', index=False, escape=False)
                    return jsonify({"status": "success", 'data_table': data_table})
                else:
                    return jsonify({"status": "error", "message": "Nothing to show!"}), 500
            else:
                return jsonify({"status": "error", "message": "please fill all required input fields"}), 500
        elif button_id == "3":
            if (from_id != 0 and to_id != 0 and id_column != ""):
                id_column = id_column.strip(" ")
                id_column = id_column.strip("\n")
                results = collection_names.find({
                    id_column: {
                        '$gte': from_id,
                        '$lte': to_id
                    }
                }, {'_id': 0})
                data = list(results)
                if len(data) > 0:
                    df2 = pd.DataFrame(data)
                    data_table = df2.to_html(classes='dataframe', index=False, escape=False)
                    return jsonify({"status": "success", 'data_table': data_table})
                else:
                    return jsonify({"status": "error", "message": "Nothing to show!"}), 500
            else:
                return jsonify({"status": "error", "message": "please fill all required input fields!"}), 500

        # handling in valid condition normally
        """
        df=pd.DataFrame()
        data_table = df.to_html(classes='dataframe', index=False, escape=False)
        return jsonify({"status": "success",'data_table': data_table})
        """
        return jsonify({"status": "error", "message": "invalid option!"}), 500
    except Exception as e:
        #print("error: ", e)
        return jsonify({"status": "error", "message": str(e)}), 500

#remove the data in the database..
@app.route('/removeTruncateDatabase', methods=['POST'])
@login_required
def removeTruncateDatabase():
    # button_id = request.args.get('id')
    # test case
    try:
        """
        p = True
        if p:
            raise FileNotFoundError("database_connection.pkl file not found.")f
        """

        if os.path.exists("./database_connection/database_connection.pkl"):
            directory = './database_connection'
            file_path = os.path.join(directory, 'database_connection.pkl')
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    conn_data = pickle.load(f)
            else:
                conn_data = {}
        client = MongoClient(
            conn_data['database_connection'],
            tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
        coll_name_temp = conn_data['source_type'] + "_" + conn_data['collection_name']
        coll_name_temp = coll_name_temp.lstrip("_")
        collection_names = client[coll_name_temp]
        
        # collection_names = client[conn_data['collection_name']]
        """
        client = MongoClient(
            "mongodb+srv://anithadevi:AnithaDevi02011998@cluster0.uyhgg.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0",
            tlsCAFile=certifi.where(), tls=True).get_database('SeoDataDatabase')
        collection_names = client['walmart_vriddhi']
        """
        data = request.json
        file_type = data.get('file_type')
        file_name = data.get('file_name')
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        button_id = data.get('id')
        date_column = data.get('date_column')
        # new added
        from_id = int(data.get('from_id'))
        to_id = int(data.get('to_id'))
        id_column = data.get('id_column')

        # print(file_type,file_name,from_date,to_date,button_id,date_column)
        deleted_count = 0
        if button_id == "1":
            if file_type == "Unique data":
                file_path = os.path.join("./unique_files", file_name)
                data = pd.read_csv(file_path)
                for x, y in zip(data['phone'], data['email']):
                    query = {
                        "$and": [
                            {"phone": x},
                            {"email": y}
                        ]
                    }
                    result = collection_names.delete_many(query)
                    if result.deleted_count == 1:
                        deleted_count += 1

                # print("method remove data execution ended")
                if deleted_count == len(data):
                    return jsonify({'status': 'success', "message": "Data Successfully removed!"})
                elif deleted_count > 0:
                    return jsonify({'status': 'success', "message": f"Only {deleted_count} Data Successfully removed!"})
                else:
                    return jsonify({'status': 'error',
                                    'message': 'Nothing to remove!'}), 500  # error rerurn
            else:
                return jsonify({'status': 'error', 'message': 'Please use Unique data!'}), 500  # error rerurn
        elif button_id == "2":
            if (from_date != "" and to_date != "" and date_column != ""):
                date_column = date_column.strip(" ")
                date_column = date_column.strip("\n")
                query = {
                    date_column: {
                        "$gte": from_date,
                        "$lte": to_date
                    }
                }
                result = collection_names.delete_many(query)
                if result.deleted_count > 0:
                    # print("successfully deleted!")
                    return jsonify({'status': 'success', "message": "Data Successfully removed!"})
                else:
                    return jsonify({'status': 'error',
                                    'message': 'Nothing to remove!'}), 500  # or another appropriate status code
            else:
                return jsonify({'status': 'error',
                                'message': 'Please fill all required input fields!'}), 500  # error return

        elif button_id == "3":
            if (from_id != 0 and to_id != 0 and id_column != ""):
                id_column = id_column.strip(" ")
                id_column = id_column.strip("\n")
                query = {
                    id_column: {
                        "$gte": from_id,
                        "$lte": to_id
                    }
                }
                result = collection_names.delete_many(query)
                if result.deleted_count > 0:
                    #print("deleted total records: ", result.deleted_count)
                    # print("successfully deleted!")
                    return jsonify({'status': 'success', "message": "Data Successfully removed!"})
                else:
                    return jsonify({'status': 'error',
                                    'message': 'Nothing to remove!'}), 500  # or another appropriate status code
            else:
                return jsonify({'status': 'error',
                                'message': 'Please fill all required input fields!'}), 500  # error rerurn

        return jsonify({'status': 'error', 'message': "Some thing problem in your operation!"})

    except Exception as e:
        return jsonify({'status': 'error',
                        'message': f'{e}'}), 500  # error rerurn

    # return Response(status=204)

#After Remove Some records/documents in the database reindex them from start
@app.route('/reindexDBid', methods=['POST'])
@login_required
def reindexDBid():
    try:
        if os.path.exists("./database_connection/database_connection.pkl"):
            directory = './database_connection'
            file_path = os.path.join(directory, 'database_connection.pkl')
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    conn_data = pickle.load(f)
            else:
                conn_data = {}
        client = MongoClient(
            conn_data['database_connection'],
            tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
        coll_name_temp = conn_data['source_type'] + "_" + conn_data['collection_name']
        coll_name_temp = coll_name_temp.lstrip("_")
        collection = client[coll_name_temp]
        if collection.count_documents({})>0:
            documents = list(collection.find({}).sort('id', 1))

            for index, doc in enumerate(documents, start=1):  # Start enumeration from 1 or 0 based on preference
                collection.update_one(
                    {'_id': doc['_id']},  # Find document by its unique _id
                    {'$set': {'id': index}}  # Set the new id value
                )
        return jsonify({"status": "success", "message": "Database re-indexed successfully!"}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

"""
                                                    Truncate DB Ends
                                                    Data Analysis Starts
"""


@app.route('/data_analysis')
@login_required
def data_analysis():
    # database connection

    try:
        # testing
        """
        p=True
        if p:
            raise FileNotFoundError("database_connection.pkl file not found.")
        """

        if os.path.exists("./database_connection/database_connection.pkl"):
            directory = './database_connection'
            file_path = os.path.join(directory, 'database_connection.pkl')
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    conn_data = pickle.load(f)
            else:
                conn_data = {}
        client = MongoClient(
            conn_data['database_connection'],
            tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
        # collection_names = client[conn_data['collection_name']]
        coll_name_temp = conn_data['source_type'] + "_" + conn_data['collection_name']
        coll_name_temp = coll_name_temp.lstrip("_")
        collection_names = client[coll_name_temp]
        """
                                                    Data Length
        """
        data_length = collection_names.count_documents({})

        if data_length > 0:
            column_names_data = list(list(collection_names.find().limit(1))[0].keys())
            required_columns = ['platform', 'business sector', 'state', 'city']
            #checking columns is present or subset of all columns fetched from database
            is_subset = set(required_columns).issubset(column_names_data)
            if is_subset:
                """
                                                    Source Count
                """
                #like instagram,facebook,...etc
                unique_values_pf = collection_names.distinct('platform') 
                #used to length of data if it contains NAN values or None values
                source_count = len(set([x for x in list(unique_values_pf) if not pd.isna(x)]))
                """
                                                    BiZ Sector Count
                """
                unique_values_bz = collection_names.distinct('business sector')
                biz_sector_count = len(set([x for x in list(unique_values_bz) if not pd.isna(x)]))
                """
                                                    Finding the Max_State
                """
                # start max records
                pipeline_state = [
                    {"$group": {"_id": "$state", "count": {"$sum": 1}}},  # Group by 'state' and count occurrences
                    {"$sort": {"count": -1}},  # Sort by count in descending order
                    {"$limit": 3}  # Limit to the top 3 result
                ]
                result = list(collection_names.aggregate(pipeline_state))
                #result = [{'_id':'Andhra pradesh,'count':74},{},{}]
                max_record_state = [x['_id'] for x in result if not pd.isna(x['_id'])][0]
                """
                                                    Finding the Max_City
                """
                pipeline_city = [
                    {"$group": {"_id": "$city", "count": {"$sum": 1}}},  # Group by 'my_field' and count occurrences
                    {"$sort": {"count": -1}},  # Sort by count in descending order
                    {"$limit": 3}  # Limit to the top 3 result
                ]
                result = list(collection_names.aggregate(pipeline_city))
                # result = [{'_id':Krishna,'count':74},{},{}]
                max_record_city = [x['_id'] for x in result if not pd.isna(x['_id'])][0]

                """
                                                    platform Count
                """
                count_fb = collection_names.count_documents({"platform": "fb"})
                count_ig = collection_names.count_documents({"platform": "ig"})
                count_website = collection_names.count_documents({"platform": "Website"})
                #print(count_fb,count_ig,count_website)
                """
                                        Handling Top 5 States alongs with counts
                """
                pipeline_pie = [
                    {
                        '$group': {
                            '_id': '$state',  # Replace 'state' with your column name
                            'count': {'$sum': 1}
                        }
                    },
                    {
                        '$sort': {'count': -1}  # Sort by count in descending order
                    },
                    {
                        '$limit': 5  # Limit the results to the top 6
                    }
                ]

                # Execute the aggregation pipeline
                state_list = []
                count_list = []
                results = collection_names.aggregate(pipeline_pie)
                for result in results:
                    if not pd.isna(result['_id']):
                        state_list.append(result['_id'])
                        count_list.append(result['count'])
                result = " ".join(conn_data['collection_name'].split("_"))
                result = result.capitalize()
                result = conn_data['source_type'] + " " + result
                result = result.strip(" ")
                
                # Example data for the cards and graphs
                card_data = {
                    'data_length': data_length,
                    'source_count': source_count,
                    'biz_sector_count': biz_sector_count,
                    'max_record_city': max_record_city,
                    'max_record_state': max_record_state,
                    'collection_name': result
                }

                # Example data for the charts
                chart_data = {
                    'bar_chart': {
                        'labels': ['Source 1', 'Source 2', 'Source 3'],
                        'data': [count_fb, count_ig, count_website]
                    },
                    'pie_chart': {
                        'labels': state_list,
                        'data': count_list
                    }
                }

                return render_template("components/data_analysis.html", card_data=card_data, chart_data=chart_data)
            else:
                return render_template("components/error.html",
                                       error_message="columns [platform, business sector, state, city] not exists in your data columns, please change your columns as listed here if exists!")
                # return jsonify({"status": "error", "message": "columns [platform, business sector, state, city] not exists in your data columns, please change your columns as listed here if exists!"}), 500
        else:
            return render_template("components/error.html",
                                   error_message="There is no data in your database!")
            # return jsonify({"status": "error","message": "There is no data in your database!"}), 500

    except Exception as e:
        # return jsonify({"status": "error", "message": str(e)}), 500
        return render_template("components/error.html",
                               error_message=f"{str(e)}")

"""
                                                    Data Analyasis Ends
                                                    Report Handling Starts


"""
# start report handling
@app.route('/report_handling')
@login_required
def report_handling():
    try:
        global ReportDataGlobal
        if ReportDataGlobal is not None:
            if len(ReportDataGlobal)>0:
                ReportDataGlobal=pd.DataFrame()
    except Exception as e:
        pass
    return render_template('components/report_generation.html')

#After Clicking "OK" button In FrontEnd Side returns column names
@app.route('/getItemsReport', methods=['POST'])
@login_required
def getItemsReport():
    global ReportDataGlobal
    try:
        file_type = request.form.get('file_type')
        file_name = request.form.get('file_name')
        if (file_type =="Select option" or file_type is None) or (file_name == "Select option" or file_name is None):
            return jsonify({"status": "error", "message": "Please select proper File type/File name!"}), 500
        #used to read file already define this function above
        ReportDataGlobal = file_reading(file_type, file_name)
        #making columns in dataframe as a list of columns
        items = ReportDataGlobal.columns.tolist()
        #sending Columns names to frontend to select the column name
        return jsonify({"status": "success", "items": items})
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"}), 500

# Report Column rename after Clicking "Rename" Button
@app.route('/reportColumnRename', methods=['POST'])
@login_required
def reportColumnRename():
    global ReportDataGlobal
    old_column = request.form.get('old_column')
    new_column = request.form.get('new_column')
    # print("old_column: ",old_column, "new_column: ",new_column)
    # print(ReportDataGlobal.columns.tolist())
    try:
        if len(ReportDataGlobal) > 0:
            if old_column in ReportDataGlobal.columns.tolist():
                ReportDataGlobal.rename(columns={old_column: new_column}, inplace=True)
                #print(ReportDataGlobal.columns.tolist())
                return jsonify({"status": "success"})
            else:
                return jsonify({"status": "error", "message": "Columns not found in dataset!"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"}), 500

#Rearrange The Column names
@app.route('/RearrangeReportColumns', methods=['POST'])
@login_required
def RearrangeReportColumns():
    # print("method called")
    global ReportDataGlobal
    try:
        data = request.json
        rearrange_column_name = data.get('list_of_all_column_names_to_rearrange')
        #making list columns selected in frontend side
        all_columns = [x.strip(" ") for x in rearrange_column_name.split(',') if len(x) > 0]
        if len(ReportDataGlobal) > 0:
            if len(ReportDataGlobal.columns.tolist()) == len(all_columns):
                if sorted(ReportDataGlobal.columns.tolist()) == sorted(all_columns):
                    ReportDataGlobal = ReportDataGlobal[all_columns]
                    return jsonify({"status": "success"})
                else:
                    return jsonify({"status": "error", "message": "Columns not matched!"}), 500
            else:
                return jsonify(
                    {"status": "error", "message": "Provided columns does not match with dataset columns!"}), 500
        else:
            return jsonify({"status": "error", "message": "Dataset is empty!"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"}), 500

#Remove the Columns
@app.route('/RemoveReportColumns', methods=['POST'])
@login_required
def RemoveReportColumns():
    # print("method called")
    global ReportDataGlobal
    try:
        data = request.json
        remove_column_name = data.get('list_of_all_column_names_to_remove')
        all_columns = [x.strip(" ") for x in remove_column_name.split(',') if len(x) > 0]
        if len(ReportDataGlobal) > 0:
            if set(all_columns).issubset(set(ReportDataGlobal.columns.tolist())):
                # print("yes its subset")
                ReportDataGlobal = ReportDataGlobal.drop(all_columns, axis=1)
                return jsonify({"status": "success"})
            else:
                return jsonify(
                    {"status": "error", "message": "Provided columns does not match with dataset columns!"}), 500
        else:
            return jsonify({"status": "error", "message": "Dataset is empty!"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"}), 500

#add new Column and value
@app.route('/addNewColumn', methods=['POST'])
@login_required
def addNewColumn():
    add_new_column = request.form.get('add_new_column')
    add_new_column_value = request.form.get('add_new_column_value')
    # print("add_new_column: ",add_new_column, "add_new_column_value: ",add_new_column_value)
    try:
        if len(ReportDataGlobal) > 0:
            # print("hi2")
            if add_new_column not in ReportDataGlobal.columns.tolist():
                ReportDataGlobal[add_new_column] = add_new_column_value
                return jsonify({"status": "success"})
            else:
                return jsonify({"status": "error", "message": "Columns name already existing in dataset!!"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"}), 500

#adding id Column to the dataframe
@app.route('/addIndexColumn', methods=['POST'])
@login_required
def addIndexColumn():
    # ("method called")
    data = request.json
    index_column_name = data.get('index_column_name')
    import_index = data.get('import_index') #having True if checked the box
    #print("index_column_name: ",index_column_name, "import_index: ",import_index)
    try:
        if import_index:
            if len(ReportDataGlobal) > 0:
                if index_column_name not in ReportDataGlobal:
                    id_list = [x for x in range(1, len(ReportDataGlobal) + 1)]
                    new_column = pd.Series(id_list, name=index_column_name)
                    ReportDataGlobal.insert(0, new_column.name, new_column)
                    # ReportDataGlobal[index_column_name]=id_list
                    return jsonify({"status": "success"})
                else:
                    return jsonify({"status": "error", "message": "Columns name already existing in dataset!!"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"}), 500

# return column names when we check at Data columns in Rename Columns
@app.route('/getReportColumns', methods=['POST'])
@login_required
def getReportColumns():
    try:
        file_type = request.form.get('file_type')
        file_name = request.form.get('file_name')
        # print("file_type: ", file_type, "file_name: ",file_name)
        if len(ReportDataGlobal) > 0:
            column_names = ReportDataGlobal.columns.tolist()
            return jsonify({"status": "success", "column_names": column_names}), 200
        elif file_type == 'Unique data':
            file_path = os.path.join("./unique_files", file_name)
            with open(file_path, 'r') as file:
                first_line = file.readline().strip()
                column_names = first_line  # Assuming the first line contains column names
            return jsonify({"status": "success", "column_names": column_names}), 200
        else:
            return jsonify({"status": "error", "message": "Please select unique file!"}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

#showdata
@app.route('/datashowReport', methods=['POST'])
@login_required
def datashowReport():
    # print("method call for export data show")
    file_type = request.form.get('file_type')
    file_name = request.form.get('file_name')
    if (file_type=="Select option" or file_type is None) or (file_name == "Select option" or file_name is None):
        return jsonify({"status": "error", "message": "Some thing problem in your data, please reload the data!"})

    if len(ReportDataGlobal) == 0 and (file_type == "" or file_type == "None") and (
            file_name == "" or file_name == "None"):
        return jsonify({"status": "error", "message": "Please select both file type and file name!"})
    if len(ReportDataGlobal) > 0:
        data_table_html = ReportDataGlobal.to_html(classes='display', index=False)
        return jsonify({"status": "success", "data_table": data_table_html})
    #if reportdata is empty showing the data in unique/duplicate folder based on file type 
    elif file_type == 'Unique data':
        file_path = os.path.join("./unique_files", file_name)
        # print("file_path:",file_path)
        try:
            df = pd.read_csv(file_path)
            data_table_html = df.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({"status": "success", "data_table": data_table_html})
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
            data_table_html = df.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({"status": "success", "data_table": data_table_html})
    elif file_type == 'Duplicate data':
        file_path = os.path.join("./duplicate_files", file_name)
        try:
            df = pd.read_csv(file_path)
            data_table_html = df.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({"status": "success", "data_table": data_table_html})
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
            data_table_html = df.to_html(classes='display', index=False)
            # return jsonify(data_table=data_table_html)
            return jsonify({"status": "success", "data_table": data_table_html})
    else:
        return jsonify({"status": "error", "message": "Some thing problem in your data, please reload the data!"})

#used to Reset Columns if changes occurs
@app.route('/getItemsReportRefresh', methods=['POST'])
@login_required
def getItemsReportRefresh():
    global ReportDataGlobal
    try:
        if len(ReportDataGlobal) > 0:
            items = ReportDataGlobal.columns.tolist()
            return jsonify({"status": "success", "items": items})
        else:
            return jsonify({"status": "error", "message": "Empty dataset!"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": f"{e}"}), 500

#Generating the Report
@app.route('/report_generation', methods=["POST"])
@login_required
def report_generation():
    try:
        # print("rg-1")
        report_type = request.form.get('report_type')
        print("report_type: ",report_type)
        file_type = request.form.get('file_type')
        file_name = request.form.get('file_name')
        index_column_name = request.form.get('index_column_name')
        check_box_access = request.form.get('set_value_checkbox') #checkbox at Rename Columns (Data columns)
        phone_no_10_digits = request.form.get('mobile_no_digit')
        impressions_count = request.form.get('report_input1')
        clicks_count = request.form.get('report_input2')
        platform_column = request.form.get('report_input3')
        rename_column = request.form.get('column_rename')
        column_list = [x.strip(" ") for x in rename_column.split(",") if len(x) > 0]
        #campaign_type = request.form.get('campaign_type')
        #campaign_column_name = request.form.get('campaign_column_name')
        # print("platform column name:",platform_column)
        # print("rg-2")
        # print(column_list)
        # print('index_column_name: ', index_column_name, 'campaign_type: ', campaign_type, 'campaign_column_name: ',
        #      campaign_column_name)

        # remove all the files available in report file
        if os.path.exists("./report_file"):
            shutil.rmtree("./report_file")

        if file_type == 'Unique data' and (file_name != "Select option" and len(file_name) > 0):
          
            file_path_new = os.path.join("./unique_files", file_name)
            data_frame = ReportDataGlobal if len(ReportDataGlobal) > 0 else pd.read_csv(file_path_new)
            #convert phone datatype into int
            data_frame['phone'] = data_frame['phone'].astype(str) 

            """
            #adding all two columns which is in report
            data_frame.insert(0, campaign_column_name, campaign_type)
            index_list=[x for x in range(1,len(data_frame)+1)]
            data_frame.insert(0, index_column_name, index_list)
        
            # print("dataframe column: ",data_frame.columns.tolist())

            # removing records which has not digits equal to 10
            # print("type of data_frame", type(data_frame), data_frame)
            # print("dataframe columns: ", data_frame.columns)
            """


            """
                                                Handling Phone Number 10 Digits
            """

            if phone_no_10_digits:
                data_frame = data_frame[data_frame['phone'].str.len() == 10]


            """
                                                Handling Platform
            """
            # print("platform_column: ",platform_column)
            if platform_column == None or platform_column == "":
                platform_column = "platform"

            if platform_column not in data_frame.columns.tolist() and report_type == "1":
                return jsonify(
                    {"status": "error", "message": f"{platform_column} column does not exists in Dataset!"}), 500

            fb_count = len(data_frame[data_frame[platform_column] == 'fb']) if platform_column in data_frame.columns.tolist() else 0
            ig_count = len(data_frame[data_frame[platform_column] == 'ig']) if platform_column in data_frame.columns.tolist() else 0

            """
                                                Creating Top DataFrame
            """

            top_dataframe = pd.DataFrame(
                {"FB Total": [fb_count], "IG Total": [ig_count], "GA Total": [0], "LI Total": [0],
                 "Impressions": [impressions_count], "Clicks": [clicks_count], "Total Leads": [len(data_frame)]})

            """
                                                Handling Rename Columns
            """

            if len(column_list) > 0 and len(data_frame.columns.tolist()) > 0 and check_box_access and len(
            column_list) == len(data_frame.columns.tolist()):
                data_frame.columns = column_list
            else:
                if report_type=="1":
                    return jsonify({"status": "error", "message": "Some thing problem in columns format!"}), 500


            """
                                                Handling Index column
            """
            if index_column_name in list(data_frame.columns):
                data_frame[index_column_name] = data_frame[index_column_name].astype(int)
                data_frame[index_column_name] = list(range(1, len(data_frame)+1))
        
            if len(data_frame)==0 and report_type=="1":
                return jsonify({"status": "error", "message": "Empty dataset, Empty report will be generated!"}), 500
    
            """
                                                Writing data from dataframe to report.xlsx file
            """

            # Define the path for the Excel file
            folder_name = "./report_file"
            if not os.path.exists(folder_name):
                os.makedirs(folder_name)
            excel_file = 'report.xlsx'
            filename = os.path.join(folder_name, excel_file)

            # Use ExcelWriter to write DataFrames to one sheet
            with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
                top_dataframe.to_excel(writer, sheet_name='Sheet1', index=False,startrow=2)  # Start the DataFrame after the heading
                data_frame.to_excel(writer, sheet_name='Sheet1', index=False,startrow=5)  # Adjust the startrow to add space
      
            # Load the workbook and select the active worksheet
            workbook = load_workbook(filename)
            worksheet = workbook.active

            """
                                            Handling Heading And its Styles
            """
            # Define formatting for the heading
            heading_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC',fill_type='solid')  # Light green background
            heading_font = Font(bold=True, size=20)  # Bold and larger font size
            heading_alignment = Alignment(horizontal='center', vertical='center')  # Center the text

            # Add heading text
            heading_cell = worksheet['A1']
            worksheet.merge_cells('A1:J1')  # Merge cells for the heading (adjust range as needed)
            heading_text = "Report of Data Collection"
            heading_cell.value = heading_text
            heading_cell.fill = heading_fill
            heading_cell.font = heading_font
            heading_cell.alignment = heading_alignment
            
            """
                                            Handling Application name and date
            """
            # application name and date
            if os.path.exists("./database_connection/database_connection.pkl"):
                directory = './database_connection'
                file_path = os.path.join(directory, 'database_connection.pkl')
                if os.path.exists(file_path):
                    with open(file_path, "rb") as f:
                        conn_data = pickle.load(f)
                else:
                    conn_data = {}
            # colletion name settings
            collection_names = conn_data['collection_name'].split("_")
            collection_names = ' '.join(collection_names)
            # new changes start
            collection_names = conn_data['source_type'] + ' ' + collection_names
            collection_names = collection_names.lstrip(' ')
            # new changes end
            name = collection_names.capitalize()
            
        
            name_date_fill = PatternFill(start_color='ffcccb', end_color='ffcccb',fill_type='solid')  # Light Salmon Pink background
            name_date_font = Font(bold=True, size=14)  # Bold and larger font size
            name_date_alignment = Alignment(horizontal='center', vertical='center')  # Center the text
            # name = "Walmart Vriddhi"
            date_extract = file_name.split('_')[-1]  # date extraction
            date_extract = date_extract[:4] + '-' + date_extract[4:6] + '-' + date_extract[6:8]  # handle date by unique file name date
            name_date_text = f"Event Name: {name}    Date: {date_extract}"
            worksheet.merge_cells('A2:J2')  # Merge cells for the heading (adjust range as needed)
            name_date_cell = worksheet['A2']
            name_date_cell.value = name_date_text
            name_date_cell.fill = name_date_fill
            name_date_cell.font = name_date_font
            name_date_cell.alignment = name_date_alignment

            # HANDLE DATAFRAME DATA:
            # Define a fill pattern for the header (yellow color in ARGB format)
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Apply the fill pattern to the header of the first DataFrame
            for cell in worksheet[3:4][0]:  # Assuming headers are in the first row
                cell.fill = header_fill

            # Apply the fill pattern to the header of the second DataFrame
            for cell in worksheet[6:7][0]:  # Assuming headers are in the 9th row after 2 rows of space
                cell.fill = header_fill

            # Save the modified Excel file
            workbook.save(filename)
            return jsonify({"status": "success", "message": "Report generated successfully!"})
        else:
            return jsonify({"status": "error", "message": "Data doesn't Exists,Please upload data file!"}), 500
        # return Response(status=204)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/report_download')
@login_required
def report_download():
    if os.path.exists('./report_file'):
        list_of_files = list(os.listdir('./report_file'))
        if len(list_of_files) > 0:
            file_path = os.path.join('./report_file', list_of_files[0])
            return send_file(file_path, as_attachment=True, download_name=list_of_files[0])
        else:
            return Response(status=204)
    else:
        return Response(status=204)

# end report handling
"""
                                                Report Ends
                                                Registration Starts
"""
# registartion page
@app.route('/registration')
@login_required
def registration():
    return render_template("pages/registration_page.html")


# if usetype is admin then only user registraion done
@app.route('/userRegistration', methods=['POST', 'GET'])
@login_required
def userRegistration():
    if request.method == 'POST':
        if session['usertype']!="admin":
            flash("You are not authorized for registering the new user!!", "error")
            return redirect(url_for('registration'))
        name = request.form.get('name')
        phone = request.form.get('phone')
        usertype = request.form.get('usertype')
        email = request.form.get('email')
        password = request.form.get('password')
        re_password = request.form.get('repassword')
        if password == re_password:
            email_regex = r'^[a-zA-Z0-9]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            phone_regex = r'^\d{10}$'
            password_regex = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"  # at least one lowercase letter,one uppercase letter, one digit, minimum length of 8 characters
            """
            if re.match(phone_regex, phone) is not None:
                print("phone valid")
            if re.match(email_regex, email) is not None:
                print("email valid")
            if re.match(password_regex, password) is not None:
                print("password valid")
            """

            if re.match(phone_regex, phone) is not None and re.match(email_regex, email) is not None and re.match(
                    password_regex, password) is not None:

                try:

                    if os.path.exists("./database_connection/database_connection.pkl"):
                        directory = './database_connection'
                        file_path = os.path.join(directory, 'database_connection.pkl')
                        if os.path.exists(file_path):
                            with open(file_path, "rb") as f:
                                conn_data = pickle.load(f)
                        else:
                            conn_data = {}
                    client = MongoClient(
                        conn_data['database_connection'],
                        tlsCAFile=certifi.where(), tls=True).get_database(conn_data['database_name'])
                    collection = client['userDetail']

                    data_doc = {"name": name, "phone": phone, "usertype": usertype, "email": email,
                                "password": password}
                    query = {
                        '$and': [
                            {"email": email},
                            {"password": password}
                        ]
                    }
                    document = collection.find_one(query)
                    if document is None:
                        result = collection.insert_one(data_doc)
                        if result.inserted_id:
                            return redirect(url_for('dashboard'))
                        else:
                            flash("Something wrong,user has not registered !!", "error")

                    else:
                        # print("part 1")
                        flash("Something goes wrong, User already exists!!", "error")
                        return redirect(url_for('registration'))
                except Exception as e:
                    flash(f"Error occured {e}!!", "error")
                    return jsonify({"status": "error", "message": f"{e}"}), 500
            else:
                # print("part 2")
                flash("Something goes wrong, your input not following pattern of input fields!!", "error")
                return redirect(url_for('registration'))
        else:
            # print("part 3")
            flash("Something goes wrong, Password not matched!!", "error")
            return redirect(url_for('registration'))
    else:
        # print("part 4")
        flash("Something goes wrong!!", "error")
        return redirect(url_for('registration'))

"""
                                                Registration Ends
                                                Logot Starts
"""
# handling logout session poppedup
@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('name', None)
    session.pop('usertype', None)
    session["is_authenticated"] = False
    return redirect(url_for('index'))

# main driver function
if __name__ == '__main__':
    app.run(debug=True)